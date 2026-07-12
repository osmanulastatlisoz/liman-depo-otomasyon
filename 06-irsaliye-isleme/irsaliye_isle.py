# -*- coding: utf-8 -*-
"""
İRSALİYE / FATURA İŞLEME YARDIMCISI
-----------------------------------
Bir klasördeki taranmış PDF irsaliyeleri işler:
  1) Her PDF'in metnini alır (PDF'te metin katmanı varsa onu; yoksa Tesseract ile OCR).
  2) Fatura/İrsaliye No + Tarih'i bulur.
  3) PDF'i  <No>.pdf  olarak yeniden adlandırır (çakışma korumalı).
  4) Özet Excel (Dosya / No / Tarih) ve tüm metinleri içeren bir .txt döker.

EN KOLAY KULLANIM:
  Dosyaya çift tıkla (veya  python irsaliye_isle.py  ).
  -> Klasör seçme penceresi açılır, klasörü seç.
  -> Ne yapacağını gösterir, "Uygula?" diye sorar. Evet dersen adlandırır.

KOMUT SATIRI (istersen):
  python irsaliye_isle.py "KLASOR_YOLU"            # kuru deneme (degistirmez)
  python irsaliye_isle.py "KLASOR_YOLU" --uygula   # gercek adlandirma
"""
import sys, os, re, glob, subprocess, tempfile
import tkinter as tk
from tkinter import filedialog, messagebox

BASLANGIC_KLASORU = os.path.join(os.path.expanduser("~"), "Desktop", "TARAMA")  # Klasor secme penceresi burada acilir (yoksa calisma klasoru)
TESS = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if not os.path.exists(TESS):          # farkli PC: PATH'teki tesseract'i kullan
    import shutil as _sh
    TESS = _sh.which("tesseract") or TESS


def pdf_text_layer(pdf_path):
    try:
        from pypdf import PdfReader
        r = PdfReader(pdf_path)
        return "\n".join((p.extract_text() or "") for p in r.pages)
    except Exception:
        return ""

def ocr_topright(pdf_path):
    """Metin katmanı yoksa: sağ-üst bilgi kutusunu render edip Tesseract'la oku.
    (Alttaki 'Fatura Açıklaması' numaralarını kapsamaz, böylece yanlış no kapılmaz.)"""
    import pypdfium2 as pdfium
    img = pdfium.PdfDocument(pdf_path)[0].render(scale=5.0).to_pil()
    w, h = img.size
    crop = img.crop((int(w*0.50), int(h*0.20), w, int(h*0.46)))
    tmp = os.path.join(tempfile.gettempdir(), "_ocr_box.png")
    crop.save(tmp)
    try:
        out = subprocess.run([TESS, tmp, "stdout", "-l", "tur+eng", "--psm", "6"],
                             capture_output=True, text=True, encoding="utf-8", errors="ignore")
        return out.stdout or ""
    except Exception:
        return ""

def ocr_full(pdf_path):
    import pypdfium2 as pdfium
    img = pdfium.PdfDocument(pdf_path)[0].render(scale=4.0).to_pil()
    tmp = os.path.join(tempfile.gettempdir(), "_ocr_full.png")
    img.save(tmp)
    try:
        out = subprocess.run([TESS, tmp, "stdout", "-l", "tur+eng", "--psm", "4"],
                             capture_output=True, text=True, encoding="utf-8", errors="ignore")
        return out.stdout or ""
    except Exception:
        return ""

NO_RE = re.compile(r"\b([A-Z]{2,3}\d{12,15})\b")
DATE_RE = re.compile(r"(\d{2}[-./]\d{2}[-./]\d{4})")

def find_no(text):
    m = re.search(r"FATURA\s*NO\s*[:\.]?\s*([A-Z0-9]{14,18})", text, re.I)
    if m:
        cand = m.group(1).upper()
        if re.fullmatch(r"[A-Z]{2,3}\d{12,15}", cand):
            return cand
    cands = NO_RE.findall(text.upper())
    return cands[0] if cands else ""

def find_date(text):
    m = DATE_RE.search(text)
    return m.group(1).replace("/", ".").replace("-", ".") if m else ""


def analyze(folder):
    """Her PDF icin karar uretir; hicbir dosyayi DEGISTIRMEZ."""
    pdfs = sorted(glob.glob(os.path.join(folder, "*.pdf")))
    results, dump = [], []
    for p in pdfs:
        name = os.path.basename(p)
        layer = pdf_text_layer(p)
        if len(layer.strip()) >= 40:
            kaynak = "metin-katmani"; no = find_no(layer); tarih = find_date(layer); text = layer
        else:
            box = ocr_topright(p); no = find_no(box); tarih = find_date(box)
            kaynak = "OCR(kutu)" if no else "OCR(kutu/BULUNAMADI)"
            text = ocr_full(p)
            if not tarih: tarih = find_date(text)
        hedef = (no + ".pdf") if no else ""
        if not no:
            action = "flag"
        elif name == hedef:
            action = "already"
        elif os.path.exists(os.path.join(folder, hedef)):
            action = "exists"
        else:
            action = "rename"
        results.append({"path": p, "name": name, "no": no, "tarih": tarih,
                        "kaynak": kaynak, "hedef": hedef, "action": action})
        dump.append(f"===== {name}  (No={no}, Tarih={tarih}) =====\n{text.strip()}\n")
    return results, dump

def apply_renames(folder, results):
    n = 0
    for r in results:
        if r["action"] == "rename":
            os.rename(r["path"], os.path.join(folder, r["hedef"])); n += 1
    return n

def _durum_text(r, applied):
    a = r["action"]
    if a == "rename":
        return ("Adlandırıldı → " + r["hedef"]) if applied else ("Adlandırılacak → " + r["hedef"])
    if a == "already": return "Zaten doğru"
    if a == "exists":  return "ATLANDI (hedef zaten var)"
    return "⚠ ELLE BAK — no bulunamadı"

def write_outputs(folder, results, dump, applied=False):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook(); ws = wb.active; ws.title = "Ozet"
        ws.append(["Dosya", "Fatura/İrsaliye No", "Tarih", "Durum"])
        for c in range(1, 5):
            ws.cell(row=1, column=c).font = Font(bold=True)
        sari = PatternFill("solid", start_color="FFFF00")
        for r in results:
            adi = r["hedef"] if (r["action"] == "rename" and applied) else r["name"]
            ws.append([adi, r["no"], r["tarih"], _durum_text(r, applied)])
            if r["action"] == "flag":                      # ELLE BAK -> satiri sariya boya
                for c in range(1, 5):
                    ws.cell(row=ws.max_row, column=c).fill = sari
        for col, w in zip("ABCD", [26, 20, 12, 34]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        wb.save(os.path.join(folder, "irsaliye_ozet.xlsx"))
    except Exception as e:
        print("Excel yazilamadi:", e)
    with open(os.path.join(folder, "irsaliye_metinleri.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(dump))

def preview_lines(results):
    out = []
    for r in results:
        a = r["action"]
        if a == "rename":   out.append(f"  {r['name']}   →   {r['hedef']}")
        elif a == "already":out.append(f"  {r['name']}   (zaten doğru)")
        elif a == "exists": out.append(f"  {r['name']}   (ATLANDI: {r['hedef']} zaten var)")
        else:               out.append(f"  {r['name']}   ⚠ ELLE BAK (no bulunamadı)")
    return out

def counts(results):
    c = lambda a: sum(1 for r in results if r["action"] == a)
    return c("rename"), c("flag"), c("already"), c("exists")


def run(folder, apply=False, interactive=False, root=None):
    results, dump = analyze(folder)
    print(f"\n{len(results)} PDF bulundu — klasör: {folder}\n")
    _et = {"rename": "adlandirilacak", "already": "zaten dogru",
           "exists": "ATLANDI (hedef var)", "flag": "!! ELLE BAK"}
    for r in results:
        durum = _et[r["action"]] + (f" -> {r['hedef']}" if r["action"] == "rename" else "")
        print(f"  {r['name']:24s} [{r['kaynak']:18s}] No={r['no'] or '-':18s} Tarih={r['tarih'] or '-':10s} {durum}")

    if not results:
        if interactive and root: messagebox.showwarning("İrsaliye İşle", f"Bu klasörde PDF yok:\n{folder}", parent=root)
        return

    ren, flag, alr, ex = counts(results)
    ozet = "\n".join(preview_lines(results)) + \
           f"\n\nAdlandırılacak: {ren}    ⚠ Elle bak: {flag}    Zaten doğru: {alr}    Atlanan: {ex}"

    do_apply = apply
    if interactive and not apply and root is not None:
        if ren > 0:
            do_apply = messagebox.askyesno(
                "İrsaliye İşle — Onay",
                ozet + "\n\nBu adlandırmalar UYGULANSIN mı?", icon="question", parent=root)
        else:
            messagebox.showinfo("İrsaliye İşle", ozet + "\n\nAdlandırılacak yeni dosya yok.", parent=root)

    if do_apply:
        n = apply_renames(folder, results)
        print(f"\n{n} dosya adlandirildi.")
    else:
        print("\n(Kuru deneme — adlandırma uygulanmadı.)")

    # Ozet Excel + metin dokumunu SON durumla yaz (Durum sutunu + ELLE BAK sari)
    write_outputs(folder, results, dump, applied=do_apply)

    if interactive and root is not None and do_apply:
        ek = f"\n\n⚠ {flag} dosya 'ELLE BAK' işaretli — adlandırılmadı.\n(irsaliye_ozet.xlsx içinde SARI satırlar.)" if flag else ""
        messagebox.showinfo("İrsaliye İşle — Bitti",
            f"{n} PDF fatura no'su ile adlandırıldı.{ek}\n\n"
            f"Aynı klasöre yazıldı:\n• irsaliye_ozet.xlsx  (Durum sütunlu)\n• irsaliye_metinleri.txt", parent=root)


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    apply = "--uygula" in sys.argv
    if args:
        run(args[0], apply=apply, interactive=False)
        return
    # Argüman yok → klasör seçme penceresi + onaylı GUI
    root = tk.Tk(); root.withdraw()
    folder = filedialog.askdirectory(
        title="İrsaliyelerin olduğu klasörü seç",
        initialdir=BASLANGIC_KLASORU if os.path.isdir(BASLANGIC_KLASORU) else os.getcwd())
    if not folder:
        print("Klasör seçilmedi, çıkılıyor."); root.destroy(); return
    run(folder, apply=False, interactive=True, root=root)
    root.destroy()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(tb)
        try:
            r = tk.Tk(); r.withdraw()
            messagebox.showerror("İrsaliye İşle — Hata", f"{e}\n\n{tb}")
            r.destroy()
        except Exception:
            pass
