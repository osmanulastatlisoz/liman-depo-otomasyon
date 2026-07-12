#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
═══════════════════════════════════════════════════════════════════════════════
  YIKAMA / TUTANAK OLUŞTURUCU
═══════════════════════════════════════════════════════════════════════════════

  KULLANIM:
      python tutanak_olustur.py

  Script seni üç şey için soracak:
      1) Tutanağın başlığı (örn: YIKAMAYA GİDECEK TUTANAĞI)
      2) İmza bloğu tipi:
            - 2_imza: Depo Sorumlusu | Genel Müdür
            - 3_imza: Depo Sorumlusu | Teslim Eden | Genel Müdür
            (Yenisini IMZA_BLOKLARI sözlüğüne ekleyebilirsin)
      3) Kısayollar ve miktarlar (örn: pm54 1, ks 3, km 1)
         - Her satıra bir tane yazabilirsin
         - Veya tek satırda "pm54 1 ks 3" gibi de yazabilirsin
         - Bitirmek için boş satır gir veya Ctrl+D / Ctrl+Z bas

  YENİ KISAYOL EKLEME:
      Aşağıdaki  KISAYOLLAR  sözlüğüne yeni satır ekle. Format:
          "KOD": ("TAM AD", "kategori", "renk", "beden")

  YENİ İMZA BLOĞU EKLEME:
      IMZA_BLOKLARI sözlüğüne yeni anahtar ekle. Örnek mevcut.

  ÇIKTI:
      Bu scriptin bulunduğu klasörde "tutanak_ciktilari/" klasörü içinde
      "[BAŞLIK] [TARİH].xlsx" adıyla dosya oluşur.

═══════════════════════════════════════════════════════════════════════════════
"""

import sys
import re
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
# AYARLAR
# ═══════════════════════════════════════════════════════════════════════════════

# Şablon dosyası bu scriptle aynı klasörde olmalı
SCRIPT_DIR    = Path(__file__).parent.resolve()
TEMPLATE_PATH = SCRIPT_DIR / "YIKAMA_TUTANAK_FORMATI.xlsx"
OUTPUT_DIR    = SCRIPT_DIR / "tutanak_ciktilari"

# Kategori sırası (üstten alta yazılacak sıra)
KATEGORI_SIRASI = [
    "kazak", "polar", "gomlek", "tshirt", "taktikal",
    "kaban", "mont", "pantolon", "tulum", "yelek",
    "baret", "ayakkabi", "cizme",
]

# Harf bedenlerinin sırası
BEDEN_SIRASI = {
    "XS": 0, "S": 1, "M": 2, "L": 3, "XL": 4,
    "2XL": 5, "3XL": 6, "4XL": 7, "5XL": 8, "6XL": 9,
}


# ═══════════════════════════════════════════════════════════════════════════════
#   ▼▼▼  İMZA BLOKLARI — YENİ İMZA TİPİ EKLEMEK / DEĞİŞTİRMEK İÇİN BURASI  ▼▼▼
# ═══════════════════════════════════════════════════════════════════════════════
#
#   Her tip bir liste: [satır1, satır2, satır3, ...]
#   Her satır da kolon bloklarının listesi:
#       (kolon_baş, kolon_son, kaç_satır_yükseklik, metin, font_boyutu)
#
#   Örnek: ("A", "F", 1, "DEPO SORUMLUSU", 12)
#     → A:F sütunlarını birleştir, 1 satır, "DEPO SORUMLUSU" yaz, font 12pt bold
#
#   Yeni tip eklemek istersen aşağıya örnekte gösterildiği gibi yeni anahtar ekle.
#
# ═══════════════════════════════════════════════════════════════════════════════

IMZA_BLOKLARI = {

    # ─── 2 İMZALI (varsayılan): Depo Sorumlusu | Genel Müdür ──────────────────
    "2_imza": [
        # (kolon_bas, kolon_son, kac_satir, metin, font_size)
        # SEPARATOR (boş satır)
        [],
        # Başlık satırı
        [("A", "F", 1, "DEPO SORUMLUSU", 12),
         ("G", "J", 1, "GENEL MÜDÜR",   12)],
        # AD SOYAD satırı (2 satır yüksekliğinde)
        [("A", "F", 2, "AD SOYAD: SEZER GÜNAY", 10),
         ("G", "J", 2, "AD SOYAD: OKAN DURAK",  10)],
        # GÖREV satırı
        [("A", "F", 2, "GÖREV: DEPO SORUMLUSU", 10),
         ("G", "J", 2, "GÖREV: GENEL MÜDÜR",    10)],
        # İMZA satırı
        [("A", "F", 2, "İMZA: ", 10),
         ("G", "J", 2, "İMZA: ", 10)],
    ],

    # ─── 3 İMZALI: Depo Sorumlusu | Teslim Eden | Genel Müdür ─────────────────
    "3_imza": [
        [],  # separator
        [("A", "D", 1, "DEPO SORUMLUSU", 11),
         ("E", "H", 1, "TESLİM EDEN",    11),
         ("I", "J", 1, "GENEL MÜDÜR",    12)],
        [("A", "D", 2, "AD SOYAD: SEZER GÜNAY", 10),
         ("E", "H", 2, "AD SOYAD: ",            10),
         ("I", "J", 2, "AD SOYAD: OKAN DURAK",  10)],
        [("A", "D", 2, "GÖREV: DEPO SORUMLUSU", 10),
         ("E", "H", 2, "GÖREV: ",               10),
         ("I", "J", 2, "GÖREV: GENEL MÜDÜR",    10)],
        [("A", "D", 2, "İMZA: ", 10),
         ("E", "H", 2, "İMZA: ", 10),
         ("I", "J", 2, "İMZA: ", 10)],
    ],

    # ─── ÖRNEK: Yeni bir tip ekle (4 imza, vs) — kullanmıyorsan silebilirsin ──
    # "4_imza": [
    #     [],
    #     [("A", "C", 1, "BAŞLIK 1", 11), ("D", "E", 1, "BAŞLIK 2", 11),
    #      ("F", "H", 1, "BAŞLIK 3", 11), ("I", "J", 1, "BAŞLIK 4", 11)],
    #     ...
    # ],
}


# ═══════════════════════════════════════════════════════════════════════════════
#   ▲▲▲  İMZA BLOKLARI BİTTİ  ▲▲▲
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
#   ▼▼▼  KISAYOLLAR — YENİ KOD EKLEMEK İÇİN BURAYI DÜZENLE  ▼▼▼
# ═══════════════════════════════════════════════════════════════════════════════
#   Format: "KOD": ("TAM AD", "kategori", "renk", "beden")
# ═══════════════════════════════════════════════════════════════════════════════

KISAYOLLAR = {

    # ─── KAZAKLAR ──────────────────────────────────────────────────────────────
    "KS":    ("POLO YAKA SWEAT S BEDEN",    "kazak", "", "S"),
    "KM":    ("POLO YAKA SWEAT M BEDEN",    "kazak", "", "M"),
    "KL":    ("POLO YAKA SWEAT L BEDEN",    "kazak", "", "L"),
    "KXL":   ("POLO YAKA SWEAT XL BEDEN",   "kazak", "", "XL"),
    "K2XL":  ("POLO YAKA SWEAT 2XL BEDEN",  "kazak", "", "2XL"),
    "K3XL":  ("POLO YAKA SWEAT 3XL BEDEN",  "kazak", "", "3XL"),
    "K4XL":  ("POLO YAKA SWEAT 4XL BEDEN",  "kazak", "", "4XL"),
    "K5XL":  ("POLO YAKA SWEAT 5XL BEDEN",  "kazak", "", "5XL"),

    # ─── POLARLAR ──────────────────────────────────────────────────────────────
    # GRİ
    "GPS":   ("GRİ POLAR S BEDEN",    "polar", "GRİ", "S"),
    "GPM":   ("GRİ POLAR M BEDEN",    "polar", "GRİ", "M"),
    "GPL":   ("GRİ POLAR L BEDEN",    "polar", "GRİ", "L"),
    "GPXL":  ("GRİ POLAR XL BEDEN",   "polar", "GRİ", "XL"),
    "GP2XL": ("GRİ POLAR 2XL BEDEN",  "polar", "GRİ", "2XL"),
    "GP3XL": ("GRİ POLAR 3XL BEDEN",  "polar", "GRİ", "3XL"),
    "GP4XL": ("GRİ POLAR 4XL BEDEN",  "polar", "GRİ", "4XL"),
    "GP5XL": ("GRİ POLAR 5XL BEDEN",  "polar", "GRİ", "5XL"),
    # HAKİ
    "HPS":   ("HAKİ POLAR S BEDEN",   "polar", "HAKİ", "S"),
    "HPM":   ("HAKİ POLAR M BEDEN",   "polar", "HAKİ", "M"),
    "HPL":   ("HAKİ POLAR L BEDEN",   "polar", "HAKİ", "L"),
    "HPXL":  ("HAKİ POLAR XL BEDEN",  "polar", "HAKİ", "XL"),
    "HP2XL": ("HAKİ POLAR 2XL BEDEN", "polar", "HAKİ", "2XL"),
    "HP3XL": ("HAKİ POLAR 3XL BEDEN", "polar", "HAKİ", "3XL"),
    "HP4XL": ("HAKİ POLAR 4XL BEDEN", "polar", "HAKİ", "4XL"),
    # KUM
    "KPS":   ("KUM POLAR S BEDEN",    "polar", "KUM", "S"),
    "KPM":   ("KUM POLAR M BEDEN",    "polar", "KUM", "M"),
    "KPL":   ("KUM POLAR L BEDEN",    "polar", "KUM", "L"),
    "KPXL":  ("KUM POLAR XL BEDEN",   "polar", "KUM", "XL"),
    "KP2XL": ("KUM POLAR 2XL BEDEN",  "polar", "KUM", "2XL"),
    # SİYAH
    "SPS":   ("SİYAH POLAR S BEDEN",   "polar", "SİYAH", "S"),
    "SPM":   ("SİYAH POLAR M BEDEN",   "polar", "SİYAH", "M"),
    "SPL":   ("SİYAH POLAR L BEDEN",   "polar", "SİYAH", "L"),
    "SPXL":  ("SİYAH POLAR XL BEDEN",  "polar", "SİYAH", "XL"),
    "SP2XL": ("SİYAH POLAR 2XL BEDEN", "polar", "SİYAH", "2XL"),
    "SP3XL": ("SİYAH POLAR 3XL BEDEN", "polar", "SİYAH", "3XL"),
    "SP4XL": ("SİYAH POLAR 4XL BEDEN", "polar", "SİYAH", "4XL"),
    "SP5XL": ("SİYAH POLAR 5XL BEDEN", "polar", "SİYAH", "5XL"),

    # ─── GÖMLEKLER ─────────────────────────────────────────────────────────────
    # GRİ
    "GGM":   ("GRİ GÖMLEK M BEDEN",    "gomlek", "GRİ", "M"),
    "GGL":   ("GRİ GÖMLEK L BEDEN",    "gomlek", "GRİ", "L"),
    "GGXL":  ("GRİ GÖMLEK XL BEDEN",   "gomlek", "GRİ", "XL"),
    "GG2XL": ("GRİ GÖMLEK 2XL BEDEN",  "gomlek", "GRİ", "2XL"),
    # HAKİ
    "HGXS":  ("HAKİ GÖMLEK XS BEDEN",  "gomlek", "HAKİ", "XS"),
    "HGS":   ("HAKİ GÖMLEK S BEDEN",   "gomlek", "HAKİ", "S"),
    "HGM":   ("HAKİ GÖMLEK M BEDEN",   "gomlek", "HAKİ", "M"),
    "HGL":   ("HAKİ GÖMLEK L BEDEN",   "gomlek", "HAKİ", "L"),
    "HGXL":  ("HAKİ GÖMLEK XL BEDEN",  "gomlek", "HAKİ", "XL"),
    "HG2XL": ("HAKİ GÖMLEK 2XL BEDEN", "gomlek", "HAKİ", "2XL"),
    "HG3XL": ("HAKİ GÖMLEK 3XL BEDEN", "gomlek", "HAKİ", "3XL"),
    # KUM
    "KGM":   ("KUM GÖMLEK M BEDEN",    "gomlek", "KUM", "M"),
    "KGL":   ("KUM GÖMLEK L BEDEN",    "gomlek", "KUM", "L"),
    "KGXL":  ("KUM GÖMLEK XL BEDEN",   "gomlek", "KUM", "XL"),
    "KG2XL": ("KUM GÖMLEK 2XL BEDEN",  "gomlek", "KUM", "2XL"),

    # ─── T-SHİRTLER ────────────────────────────────────────────────────────────
    "STXS":  ("SİYAH T-SHİRT XS BEDEN",  "tshirt", "SİYAH", "XS"),
    "STS":   ("SİYAH T-SHİRT S BEDEN",   "tshirt", "SİYAH", "S"),
    "STM":   ("SİYAH T-SHİRT M BEDEN",   "tshirt", "SİYAH", "M"),
    "STL":   ("SİYAH T-SHİRT L BEDEN",   "tshirt", "SİYAH", "L"),
    "STXL":  ("SİYAH T-SHİRT XL BEDEN",  "tshirt", "SİYAH", "XL"),
    "ST2XL": ("SİYAH T-SHİRT 2XL BEDEN", "tshirt", "SİYAH", "2XL"),
    "ST3XL": ("SİYAH T-SHİRT 3XL BEDEN", "tshirt", "SİYAH", "3XL"),
    "ST4XL": ("SİYAH T-SHİRT 4XL BEDEN", "tshirt", "SİYAH", "4XL"),

    # ─── TAKTİKAL PANTOLONLAR ──────────────────────────────────────────────────
    "TPXS":  ("TAKTİKAL PANTOLON XS BEDEN",  "taktikal", "", "XS"),
    "TPS":   ("TAKTİKAL PANTOLON S BEDEN",   "taktikal", "", "S"),
    "TPM":   ("TAKTİKAL PANTOLON M BEDEN",   "taktikal", "", "M"),
    "TPL":   ("TAKTİKAL PANTOLON L BEDEN",   "taktikal", "", "L"),
    "TPXL":  ("TAKTİKAL PANTOLON XL BEDEN",  "taktikal", "", "XL"),
    "TP2XL": ("TAKTİKAL PANTOLON 2XL BEDEN", "taktikal", "", "2XL"),
    "TP3XL": ("TAKTİKAL PANTOLON 3XL BEDEN", "taktikal", "", "3XL"),
    "TP4XL": ("TAKTİKAL PANTOLON 4XL BEDEN", "taktikal", "", "4XL"),
    "TP5XL": ("TAKTİKAL PANTOLON 5XL BEDEN", "taktikal", "", "5XL"),

    # ─── KABANLAR (M ön ekiyle başlar) ─────────────────────────────────────────
    # GRİ (MG..)
    "MGS":   ("KABAN GRİ S BEDEN",    "kaban", "GRİ", "S"),
    "MGM":   ("KABAN GRİ M BEDEN",    "kaban", "GRİ", "M"),
    "MGL":   ("KABAN GRİ L BEDEN",    "kaban", "GRİ", "L"),
    "MGXL":  ("KABAN GRİ XL BEDEN",   "kaban", "GRİ", "XL"),
    "MG2XL": ("KABAN GRİ 2XL BEDEN",  "kaban", "GRİ", "2XL"),
    "MG3XL": ("KABAN GRİ 3XL BEDEN",  "kaban", "GRİ", "3XL"),
    "MG4XL": ("KABAN GRİ 4XL BEDEN",  "kaban", "GRİ", "4XL"),
    # KIRMIZI (MK..)
    "MKS":   ("KABAN KIRMIZI S BEDEN",   "kaban", "KIRMIZI", "S"),
    "MKM":   ("KABAN KIRMIZI M BEDEN",   "kaban", "KIRMIZI", "M"),
    "MKL":   ("KABAN KIRMIZI L BEDEN",   "kaban", "KIRMIZI", "L"),
    "MKXL":  ("KABAN KIRMIZI XL BEDEN",  "kaban", "KIRMIZI", "XL"),
    "MK2XL": ("KABAN KIRMIZI 2XL BEDEN", "kaban", "KIRMIZI", "2XL"),
    "MK3XL": ("KABAN KIRMIZI 3XL BEDEN", "kaban", "KIRMIZI", "3XL"),
    "MK4XL": ("KABAN KIRMIZI 4XL BEDEN", "kaban", "KIRMIZI", "4XL"),
    "MK5XL": ("KABAN KIRMIZI 5XL BEDEN", "kaban", "KIRMIZI", "5XL"),
    "MK6XL": ("KABAN KIRMIZI 6XL BEDEN", "kaban", "KIRMIZI", "6XL"),
    # LACİVERT (ML..)
    "MLS":   ("KABAN LACİVERT S BEDEN",   "kaban", "LACİVERT", "S"),
    "MLM":   ("KABAN LACİVERT M BEDEN",   "kaban", "LACİVERT", "M"),
    "MLL":   ("KABAN LACİVERT L BEDEN",   "kaban", "LACİVERT", "L"),
    "MLXL":  ("KABAN LACİVERT XL BEDEN",  "kaban", "LACİVERT", "XL"),
    "ML2XL": ("KABAN LACİVERT 2XL BEDEN", "kaban", "LACİVERT", "2XL"),
    "ML3XL": ("KABAN LACİVERT 3XL BEDEN", "kaban", "LACİVERT", "3XL"),
    "ML4XL": ("KABAN LACİVERT 4XL BEDEN", "kaban", "LACİVERT", "4XL"),
    "ML5XL": ("KABAN LACİVERT 5XL BEDEN", "kaban", "LACİVERT", "5XL"),
    "ML6XL": ("KABAN LACİVERT 6XL BEDEN", "kaban", "LACİVERT", "6XL"),
    # MAVİ (MM..)
    "MMS":   ("KABAN MAVİ S BEDEN",   "kaban", "MAVİ", "S"),
    "MMM":   ("KABAN MAVİ M BEDEN",   "kaban", "MAVİ", "M"),
    "MML":   ("KABAN MAVİ L BEDEN",   "kaban", "MAVİ", "L"),
    "MMXL":  ("KABAN MAVİ XL BEDEN",  "kaban", "MAVİ", "XL"),
    "MM2XL": ("KABAN MAVİ 2XL BEDEN", "kaban", "MAVİ", "2XL"),
    "MM3XL": ("KABAN MAVİ 3XL BEDEN", "kaban", "MAVİ", "3XL"),
    "MM4XL": ("KABAN MAVİ 4XL BEDEN", "kaban", "MAVİ", "4XL"),
    "MM5XL": ("KABAN MAVİ 5XL BEDEN", "kaban", "MAVİ", "5XL"),
    # TURUNCU (MT..)
    "MTS":   ("KABAN TURUNCU S BEDEN",   "kaban", "TURUNCU", "S"),
    "MTM":   ("KABAN TURUNCU M BEDEN",   "kaban", "TURUNCU", "M"),
    "MTL":   ("KABAN TURUNCU L BEDEN",   "kaban", "TURUNCU", "L"),
    "MTXL":  ("KABAN TURUNCU XL BEDEN",  "kaban", "TURUNCU", "XL"),
    "MT2XL": ("KABAN TURUNCU 2XL BEDEN", "kaban", "TURUNCU", "2XL"),
    "MT3XL": ("KABAN TURUNCU 3XL BEDEN", "kaban", "TURUNCU", "3XL"),
    "MT4XL": ("KABAN TURUNCU 4XL BEDEN", "kaban", "TURUNCU", "4XL"),
    "MT5XL": ("KABAN TURUNCU 5XL BEDEN", "kaban", "TURUNCU", "5XL"),
    # YEŞİL (MY..)
    "MYS":   ("KABAN YEŞİL S BEDEN",   "kaban", "YEŞİL", "S"),
    "MYM":   ("KABAN YEŞİL M BEDEN",   "kaban", "YEŞİL", "M"),
    "MYL":   ("KABAN YEŞİL L BEDEN",   "kaban", "YEŞİL", "L"),
    "MYXL":  ("KABAN YEŞİL XL BEDEN",  "kaban", "YEŞİL", "XL"),
    "MY2XL": ("KABAN YEŞİL 2XL BEDEN", "kaban", "YEŞİL", "2XL"),
    "MY3XL": ("KABAN YEŞİL 3XL BEDEN", "kaban", "YEŞİL", "3XL"),
    "MY4XL": ("KABAN YEŞİL 4XL BEDEN", "kaban", "YEŞİL", "4XL"),
    "MY5XL": ("KABAN YEŞİL 5XL BEDEN", "kaban", "YEŞİL", "5XL"),
    "MY6XL": ("KABAN YEŞİL 6XL BEDEN", "kaban", "YEŞİL", "6XL"),

    # ─── MONTLAR ───────────────────────────────────────────────────────────────
    # OXFORD KABAN (MO..)
    "MOS":   ("OXFORD KABAN S BEDEN",    "mont", "OXFORD", "S"),
    "MOM":   ("OXFORD KABAN M BEDEN",    "mont", "OXFORD", "M"),
    "MOL":   ("OXFORD KABAN L BEDEN",    "mont", "OXFORD", "L"),
    "MOXL":  ("OXFORD KABAN XL BEDEN",   "mont", "OXFORD", "XL"),
    "MO2XL": ("OXFORD KABAN 2XL BEDEN",  "mont", "OXFORD", "2XL"),
    "MO3XL": ("OXFORD KABAN 3XL BEDEN",  "mont", "OXFORD", "3XL"),
    "MO4XL": ("OXFORD KABAN 4XL BEDEN",  "mont", "OXFORD", "4XL"),
    "MO5XL": ("OXFORD KABAN 5XL BEDEN",  "mont", "OXFORD", "5XL"),
    # SARI MONT (SAM..)
    "SAMS":   ("SARI MONT S BEDEN",    "mont", "SARI", "S"),
    "SAMM":   ("SARI MONT M BEDEN",    "mont", "SARI", "M"),
    "SAML":   ("SARI MONT L BEDEN",    "mont", "SARI", "L"),
    "SAMXL":  ("SARI MONT XL BEDEN",   "mont", "SARI", "XL"),
    "SAM2XL": ("SARI MONT 2XL BEDEN",  "mont", "SARI", "2XL"),
    "SAM3XL": ("SARI MONT 3XL BEDEN",  "mont", "SARI", "3XL"),
    "SAM4XL": ("SARI MONT 4XL BEDEN",  "mont", "SARI", "4XL"),
    "SAM5XL": ("SARI MONT 5XL BEDEN",  "mont", "SARI", "5XL"),
    # SİYAH MONT (SM..)
    "SMS":   ("SİYAH MONT S BEDEN",    "mont", "SİYAH", "S"),
    "SMM":   ("SİYAH MONT M BEDEN",    "mont", "SİYAH", "M"),
    "SML":   ("SİYAH MONT L BEDEN",    "mont", "SİYAH", "L"),
    "SMXL":  ("SİYAH MONT XL BEDEN",   "mont", "SİYAH", "XL"),
    "SM2XL": ("SİYAH MONT 2XL BEDEN",  "mont", "SİYAH", "2XL"),
    "SM3XL": ("SİYAH MONT 3XL BEDEN",  "mont", "SİYAH", "3XL"),
    "SM4XL": ("SİYAH MONT 4XL BEDEN",  "mont", "SİYAH", "4XL"),
    "SM5XL": ("SİYAH MONT 5XL BEDEN",  "mont", "SİYAH", "5XL"),

    # ─── PANTOLONLAR (renk + numara) ───────────────────────────────────────────
    # GRİ (PG..)
    **{f"PG{n}": (f"PANTOLON GRİ {n} BEDEN", "pantolon", "GRİ", str(n))
       for n in range(48, 68, 2)},
    # KIRMIZI (PK..)
    **{f"PK{n}": (f"PANTOLON KIRMIZI {n} BEDEN", "pantolon", "KIRMIZI", str(n))
       for n in range(48, 66, 2)},
    # LACİVERT (PL..)
    **{f"PL{n}": (f"PANTOLON LACİVERT {n} BEDEN", "pantolon", "LACİVERT", str(n))
       for n in range(48, 72, 2)},
    # MAVİ (PM..)
    **{f"PM{n}": (f"PANTOLON MAVİ {n} BEDEN", "pantolon", "MAVİ", str(n))
       for n in range(48, 66, 2)},
    # TURUNCU (PT..)
    **{f"PT{n}": (f"PANTOLON TURUNCU {n} BEDEN", "pantolon", "TURUNCU", str(n))
       for n in range(48, 68, 2)},
    # YEŞİL (PY..)
    **{f"PY{n}": (f"PANTOLON YEŞİL {n} BEDEN", "pantolon", "YEŞİL", str(n))
       for n in range(48, 68, 2)},

    # ─── TULUMLAR ──────────────────────────────────────────────────────────────
    "KTS":   ("KIRMIZI TULUM S BEDEN",   "tulum", "KIRMIZI", "S"),
    "KTM":   ("KIRMIZI TULUM M BEDEN",   "tulum", "KIRMIZI", "M"),
    "KTL":   ("KIRMIZI TULUM L BEDEN",   "tulum", "KIRMIZI", "L"),
    "KTXL":  ("KIRMIZI TULUM XL BEDEN",  "tulum", "KIRMIZI", "XL"),
    "KT2XL": ("KIRMIZI TULUM 2XL BEDEN", "tulum", "KIRMIZI", "2XL"),
    "KT3XL": ("KIRMIZI TULUM 3XL BEDEN", "tulum", "KIRMIZI", "3XL"),
    "KT4XL": ("KIRMIZI TULUM 4XL BEDEN", "tulum", "KIRMIZI", "4XL"),
    "KT5XL": ("KIRMIZI TULUM 5XL BEDEN", "tulum", "KIRMIZI", "5XL"),

    # ─── YELEKLER ──────────────────────────────────────────────────────────────
    "YS":    ("YELEK SARI S BEDEN",    "yelek", "SARI", "S"),
    "YM":    ("YELEK SARI M BEDEN",    "yelek", "SARI", "M"),
    "YL":    ("YELEK SARI L BEDEN",    "yelek", "SARI", "L"),
    "YXL":   ("YELEK SARI XL BEDEN",   "yelek", "SARI", "XL"),
    "Y2XL":  ("YELEK SARI 2XL BEDEN",  "yelek", "SARI", "2XL"),
    "Y3XL":  ("YELEK SARI 3XL BEDEN",  "yelek", "SARI", "3XL"),
    "Y4XL":  ("YELEK SARI 4XL BEDEN",  "yelek", "SARI", "4XL"),
    "Y5XL":  ("YELEK SARI 5XL BEDEN",  "yelek", "SARI", "5XL"),

    # ─── BARETLER ──────────────────────────────────────────────────────────────
    "BB":    ("BARET BEYAZ KARAM ENDÜSTRİYEL KB505",   "baret", "BEYAZ",   "-"),
    "BK":    ("BARET KIRMIZI KARAM ENDÜSTRİYEL KB505", "baret", "KIRMIZI", "-"),
    "BM":    ("BARET MAVİ KARAM ENDÜSTRİYEL KB505",    "baret", "MAVİ",    "-"),
    "BT":    ("BARET TURUNCU KARAM ENDÜSTRİYEL KB505", "baret", "TURUNCU", "-"),

    # ─── AYAKKABILAR (BO..) ─────────────────────────────────────────────────────
    **{f"BO{n}": (f"AYAKKABI {n} STARLINE 9040B-83", "ayakkabi", "", str(n))
       for n in range(37, 48)},
    # ─── ÇİZMELER (Ç..) ─────────────────────────────────────────────────────────
    **{f"Ç{n}":  (f"ÇİZME {n} STARLINE 9910-S5-YEŞİL", "cizme", "", str(n))
       for n in range(37, 48)},
    **{f"C{n}":  (f"ÇİZME {n} STARLINE 9910-S5-YEŞİL", "cizme", "", str(n))
       for n in range(37, 48)},  # Ç yerine C yazılırsa da bulsun
}

# ═══════════════════════════════════════════════════════════════════════════════
#   ▲▲▲  KISAYOLLAR BİTTİ — AŞAĞIDAKİ KODA DOKUNMAYA GEREK YOK  ▲▲▲
# ═══════════════════════════════════════════════════════════════════════════════


def beden_sirasi_anahtari(beden: str) -> tuple:
    """Beden sıralaması için anahtar üretir (XS<S<M<L<XL<2XL<... ve sayısal)."""
    b = beden.upper().strip()
    if b in BEDEN_SIRASI:
        return (0, BEDEN_SIRASI[b])
    try:
        return (1, int(b))
    except ValueError:
        return (2, b)


def kategori_sirasi_anahtari(kat: str) -> int:
    try:
        return KATEGORI_SIRASI.index(kat)
    except ValueError:
        return 999


def kisayollari_coz(metin: str) -> dict:
    """
    Kullanıcının yazdığı metni 'KOD MİKTAR' çiftlerine böl, sözlüğe çevir.
    Aynı kodlar otomatik toplanır.
    Virgül, satır sonu, fazla boşluk hepsi ayraç sayılır.
    Tanınmayan kodlar uyarı listesine eklenir.
    """
    # Boşluk, virgül, noktalı virgül, tab, yeni satır — hepsi ayraç
    tokens = re.split(r"[\s,;]+", metin.strip())
    tokens = [t for t in tokens if t]

    sonuc = {}      # kod (upper) -> toplam miktar
    bilinmeyenler = []

    i = 0
    while i < len(tokens):
        tok = tokens[i]
        # Sayı mı?
        if tok.isdigit():
            # önceki kod yoksa atla
            i += 1
            continue
        # Kod
        kod = tok.upper()
        # Sonraki token sayı mı?
        miktar = 1
        if i + 1 < len(tokens) and tokens[i + 1].isdigit():
            miktar = int(tokens[i + 1])
            i += 2
        else:
            i += 1

        if kod not in KISAYOLLAR:
            bilinmeyenler.append(kod)
            continue

        sonuc[kod] = sonuc.get(kod, 0) + miktar

    return sonuc, bilinmeyenler


def sirala(toplam: dict) -> list:
    """
    Kod->miktar sözlüğünü kategori/renk/beden sırasına sokup
    [(tam_ad, miktar), ...] listesi döner.
    """
    kayitlar = []
    for kod, miktar in toplam.items():
        tam_ad, kat, renk, beden = KISAYOLLAR[kod]
        kayitlar.append((
            kategori_sirasi_anahtari(kat),
            renk,
            beden_sirasi_anahtari(beden),
            tam_ad,
            miktar,
        ))
    kayitlar.sort(key=lambda r: (r[0], r[1], r[2]))
    return [(r[3], r[4]) for r in kayitlar]


def _stil_kopyala(kaynak_hucre):
    """Bir hücrenin stil objelerini kopyalayıp dict döner."""
    if not kaynak_hucre.has_style:
        return None
    return {
        "font":          copy(kaynak_hucre.font),
        "border":        copy(kaynak_hucre.border),
        "fill":          copy(kaynak_hucre.fill),
        "alignment":     copy(kaynak_hucre.alignment),
        "number_format": kaynak_hucre.number_format,
        "protection":    copy(kaynak_hucre.protection),
    }


def _stil_uygula(hucre, stil):
    """Dict halindeki stili hücreye uygular."""
    if stil is None:
        return
    hucre.font          = copy(stil["font"])
    hucre.border        = copy(stil["border"])
    hucre.fill          = copy(stil["fill"])
    hucre.alignment     = copy(stil["alignment"])
    hucre.number_format = stil["number_format"]
    hucre.protection    = copy(stil["protection"])


def imza_yaz(ws, baslangic_satiri: int, imza_tipi: str):
    """
    IMZA_BLOKLARI[imza_tipi] tanımına göre imza bloğunu çizer.
    baslangic_satiri: separator (boş) satırın yazılacağı satır numarası.
    """
    if imza_tipi not in IMZA_BLOKLARI:
        raise ValueError(
            f"Bilinmeyen imza tipi: '{imza_tipi}'. "
            f"Mevcut tipler: {list(IMZA_BLOKLARI.keys())}"
        )

    from openpyxl.utils import column_index_from_string

    bloklar = IMZA_BLOKLARI[imza_tipi]
    su_anki_satir = baslangic_satiri

    # Üst kenarlık için ince çizgi (görsel ayraç istersen)
    # Şu an separator satırı tamamen boş; ihtiyaç olursa burada border ekle.

    for satir_def in bloklar:
        if not satir_def:
            # Boş ayraç satırı
            ws.row_dimensions[su_anki_satir].height = 14
            su_anki_satir += 1
            continue

        # Bu satırın yüksekliğini belirleyen ilk hücrenin kac_satir değerini al
        kac_satir = satir_def[0][2]

        for kol_bas, kol_son, _, metin, font_size in satir_def:
            c_bas = column_index_from_string(kol_bas)
            c_son = column_index_from_string(kol_son)

            # Birleştir
            ws.merge_cells(
                start_row=su_anki_satir, start_column=c_bas,
                end_row=su_anki_satir + kac_satir - 1, end_column=c_son,
            )

            # Sol üst hücreye değer ve stil
            hucre = ws.cell(row=su_anki_satir, column=c_bas)
            hucre.value = metin
            hucre.font = Font(name="Calibri", size=font_size, bold=True)
            # Hizalama: başlık satırlarında merkez, diğerlerinde sola dayalı
            if kac_satir == 1:
                hucre.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                hucre.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Sonraki satıra geç
        su_anki_satir += kac_satir


def tutanak_olustur(baslik: str, kalemler: list, cikti_yolu: Path,
                    imza_tipi: str = "2_imza"):
    """
    Şablonu yükler, başlığı/tarihi/satırları yazar, imza bloğunu çizer, kaydeder.
    kalemler: [(tam_ad, miktar), ...]  (sıralanmış halde)
    imza_tipi: "2_imza" veya "3_imza"  (IMZA_BLOKLARI'ndaki anahtarlar)
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Şablon bulunamadı:\n  {TEMPLATE_PATH}\n"
            f"YIKAMA_TUTANAK_FORMATI.xlsx dosyasının scriptle aynı klasörde "
            f"olduğundan emin ol."
        )

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ─── Başlık ve tarih ─────────────────────────────────────────────────────
    bugun = datetime.now().strftime("%d.%m.%Y")
    ws["A4"] = baslik.upper()
    ws["I1"] = f"   TARİH: {bugun}"

    N = max(len(kalemler), 1)
    DATA_BAS = 7        # veri ilk satırı
    IMZA_BAS = 33       # imza bloğu (separator dahil) şablonda nerede
    IMZA_SON = 40       # imza bloğu son satırı

    # ─── 1) Veri satırı stilini yakala (satır 7'den) ────────────────────────
    veri_stil_a = _stil_kopyala(ws.cell(row=DATA_BAS, column=1))
    veri_stil_i = _stil_kopyala(ws.cell(row=DATA_BAS, column=9))
    veri_satir_yuksekligi = ws.row_dimensions[DATA_BAS].height or 27

    # ─── 2) Mevcut imza bloğunu temizle (yenisi yazılacak) ──────────────────
    # (Şablonda satır 33-40 arası 2-imzalı blok var; programatik olarak yeniden çizilecek)

    # ─── 3) 7. satırdan itibaren her şeyi temizle ────────────────────────────
    # Önce birleştirmeleri kaldır
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= DATA_BAS:
            ws.unmerge_cells(str(mr))
    # Sonra hücre değerleri ve stilleri
    son_temizleme_satiri = max(IMZA_SON, DATA_BAS + N + 10)
    for row in ws.iter_rows(min_row=DATA_BAS, max_row=son_temizleme_satiri):
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.border = Border()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
    # Yükseklikleri sıfırla
    for r in range(DATA_BAS, son_temizleme_satiri + 1):
        ws.row_dimensions[r].height = None

    # ─── 4) Veri satırlarını yaz ─────────────────────────────────────────────
    kucult = N > 26  # çok kalem varsa yazıyı küçült

    for idx, (ad, miktar) in enumerate(kalemler):
        r = DATA_BAS + idx
        # A:H sütunlarına stil
        for col in range(1, 9):
            _stil_uygula(ws.cell(row=r, column=col), veri_stil_a)
        # I:J sütunlarına stil
        for col in range(9, 11):
            _stil_uygula(ws.cell(row=r, column=col), veri_stil_i)
        # Birleştirmeler
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
        # Değerler
        ws.cell(row=r, column=1).value = ad
        ws.cell(row=r, column=9).value = miktar
        # Boyut
        if kucult:
            ws.row_dimensions[r].height = 18
            ca = ws.cell(row=r, column=1)
            ci = ws.cell(row=r, column=9)
            ca.font = Font(name=ca.font.name or "Calibri", size=11,
                           bold=ca.font.bold, color=ca.font.color)
            ci.font = Font(name=ci.font.name or "Calibri", size=11,
                           bold=ci.font.bold, color=ci.font.color)
        else:
            ws.row_dimensions[r].height = veri_satir_yuksekligi

    # Eğer hiç kalem yoksa en az 1 boş satır olsun
    if len(kalemler) == 0:
        for col in range(1, 9):
            _stil_uygula(ws.cell(row=DATA_BAS, column=col), veri_stil_a)
        for col in range(9, 11):
            _stil_uygula(ws.cell(row=DATA_BAS, column=col), veri_stil_i)
        ws.merge_cells(start_row=DATA_BAS, start_column=1,
                       end_row=DATA_BAS, end_column=8)
        ws.merge_cells(start_row=DATA_BAS, start_column=9,
                       end_row=DATA_BAS, end_column=10)
        ws.row_dimensions[DATA_BAS].height = veri_satir_yuksekligi

    # ─── 5) İmza bloğunu yeniden inşa et ─────────────────────────────────────
    yeni_imza_bas = DATA_BAS + N  # ilk imza satırı (separator)

    imza_yaz(ws, yeni_imza_bas, imza_tipi)

    # ─── 6) Kaydet ───────────────────────────────────────────────────────────
    cikti_yolu.parent.mkdir(parents=True, exist_ok=True)
    wb.save(cikti_yolu)


def metin_oku(prompt: str) -> str:
    """Çok satırlı girdi oku — boş satır veya EOF ile biter."""
    print(prompt)
    satirlar = []
    try:
        while True:
            s = input()
            if s.strip() == "" and satirlar:
                break
            if s.strip() == "" and not satirlar:
                continue
            satirlar.append(s)
    except EOFError:
        pass
    return "\n".join(satirlar)


def slugify(s: str) -> str:
    """Dosya adı için güvenli karakterler bırak."""
    s = s.strip()
    s = re.sub(r"[\\/:*?\"<>|]", "_", s)
    s = re.sub(r"\s+", " ", s)
    return s[:120]


def main():
    print("═" * 70)
    print("  YIKAMA TUTANAĞI OLUŞTURUCU")
    print("═" * 70)
    print()

    # 1) Başlık
    baslik = input("Tutanak başlığı (Enter → 'YIKAMAYA GİDECEK TUTANAĞI'): ").strip()
    if not baslik:
        baslik = "YIKAMAYA GİDECEK TUTANAĞI"

    print()

    # 2) İmza tipi
    print("İmza bloğu tipini seç:")
    tipler = list(IMZA_BLOKLARI.keys())
    for i, t in enumerate(tipler, 1):
        aciklama = {
            "2_imza": "Depo Sorumlusu | Genel Müdür",
            "3_imza": "Depo Sorumlusu | Teslim Eden | Genel Müdür",
        }.get(t, t)
        print(f"  {i}) {t}  ({aciklama})")
    secim = input(f"Seçim [1-{len(tipler)}, Enter → 1]: ").strip()
    if not secim:
        imza_tipi = tipler[0]
    else:
        try:
            imza_tipi = tipler[int(secim) - 1]
        except (ValueError, IndexError):
            print(f"⚠  Geçersiz seçim, '{tipler[0]}' kullanılıyor.")
            imza_tipi = tipler[0]

    print()

    # 2) Kısayollar
    metin = metin_oku(
        "Kısayolları ve miktarları gir.\n"
        "  Örnek:  pm54 1   ks 3   km 1\n"
        "  Veya satır satır yaz. Bitirmek için boş satır gir.\n"
        "─" * 70
    )

    if not metin.strip():
        print("⚠  Hiçbir şey girilmedi, çıkılıyor.")
        sys.exit(1)

    # 3) Çöz
    toplam, bilinmeyenler = kisayollari_coz(metin)

    if bilinmeyenler:
        print()
        print("⚠  Tanınmayan kodlar (atlandı):", ", ".join(sorted(set(bilinmeyenler))))

    if not toplam:
        print("⚠  Geçerli hiçbir kod bulunamadı.")
        sys.exit(1)

    # 4) Sırala
    kalemler = sirala(toplam)

    # 5) Önizleme
    print()
    print("─" * 70)
    print(f"  {baslik}   ({len(kalemler)} kalem, toplam {sum(m for _,m in kalemler)} adet)")
    print("─" * 70)
    for ad, miktar in kalemler:
        print(f"  {ad:<50} {miktar:>5}")
    print("─" * 70)

    onay = input("Bu şekilde oluşturulsun mu? [E/h]: ").strip().lower()
    if onay and onay not in ("e", "evet", "y", "yes"):
        print("İptal edildi.")
        sys.exit(0)

    # 6) Dosya yaz
    bugun = datetime.now().strftime("%d.%m.%Y")
    dosya_adi = f"{slugify(baslik)} {bugun}.xlsx"
    cikti_yolu = OUTPUT_DIR / dosya_adi
    tutanak_olustur(baslik, kalemler, cikti_yolu, imza_tipi=imza_tipi)

    print()
    print("✓ Tutanak oluşturuldu:")
    print(f"  {cikti_yolu}")


if __name__ == "__main__":
    main()