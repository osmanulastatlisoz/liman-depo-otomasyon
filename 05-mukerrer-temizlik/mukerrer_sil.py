# -*- coding: utf-8 -*-
"""MUKERRER_KAYIT_RAPORU.xlsx içinde SİL? sütununa E yazılan grupları
2026 DEPO STOK.xlsm'den siler (her grubun İLK kaydı kalır, fazlası gider).

Excel'i HİÇ çalıştırmaz; dosyanın içindeki XML'i doğrudan düzenler. Böylece
makrolar, butonlar ve diğer tablolar bire bir korunur ve Excel lisansı
aktif olmasa bile çalışır. Çalışmadan önce otomatik tarihli yedek alır,
silmeden önce her hedef satırın içeriğini doğrular."""
import os
import json
import shutil
import sys
import warnings
from collections import defaultdict
from datetime import datetime

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _satir_sil_xml import satirlari_sil

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8")

AYAR_JSON = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
                         "DepoAraclari", "yollar.json")


def _kaynak_bul():
    """2026 DEPO STOK.xlsm: kayıtlı ayar -> masaüstü -> kullanıcıya sor
    (mukerrer_tespit.py ile aynı ayar anahtarını paylaşır)."""
    try:
        with open(AYAR_JSON, encoding="utf-8") as f:
            kayit = json.load(f)
    except Exception:
        kayit = {}
    yol = kayit.get("depo_stok_2026")
    if yol and os.path.exists(yol):
        return yol
    varsayilan = os.path.join(os.path.expanduser("~"), "Desktop", "2026 DEPO STOK.xlsm")
    if os.path.exists(varsayilan):
        return varsayilan
    import tkinter as tk
    from tkinter import filedialog
    kok = tk.Tk(); kok.withdraw()
    yol = filedialog.askopenfilename(title="2026 DEPO STOK.xlsm dosyasını seç",
                                     filetypes=[("Excel", "*.xlsm *.xlsx")])
    kok.destroy()
    if not yol:
        sys.exit("Dosya seçilmedi, çıkılıyor.")
    kayit["depo_stok_2026"] = yol
    try:
        os.makedirs(os.path.dirname(AYAR_JSON), exist_ok=True)
        with open(AYAR_JSON, "w", encoding="utf-8") as f:
            json.dump(kayit, f, ensure_ascii=False, indent=1)
    except OSError:
        pass
    return yol


DOSYA = _kaynak_bul()
RAPOR = os.path.join(os.path.dirname(DOSYA), "MUKERRER_KAYIT_RAPORU.xlsx")

# 1) Rapordan işaretli grupları oku
try:
    rwb = openpyxl.load_workbook(RAPOR, read_only=True, data_only=True)
except FileNotFoundError:
    print("HATA: Rapor bulunamadı. Önce '1_TARAMA YAP' ı çalıştırın.")
    sys.exit(1)
if "MÜKERRER KAYITLAR" not in rwb.sheetnames:
    print("HATA: Rapor eski biçimde. Önce '1_TARAMA YAP' ı yeniden çalıştırın.")
    sys.exit(1)
rws = rwb["MÜKERRER KAYITLAR"]
isaretli = set()
for row in rws.iter_rows(min_row=2, values_only=True):
    sil, grup, tarih_s, personel, stok = (row + (None,) * 9)[:5]
    if sil is None or not str(sil).strip():
        continue
    if not (tarih_s and personel and stok):
        continue
    try:
        t = datetime.strptime(str(tarih_s).strip(), "%d.%m.%Y").date()
    except ValueError:
        print(f"UYARI: '{tarih_s}' tarihi okunamadı, bu satır atlandı.")
        continue
    isaretli.add((str(personel).strip().upper(), str(stok).strip().upper(), t))
rwb.close()

if not isaretli:
    print("Raporda SİL? sütununa E yazılmış hiçbir kayıt yok.")
    print("Önce raporu açıp silinecekleri işaretleyin, kaydedin, sonra tekrar çalıştırın.")
    sys.exit(0)
print(f"Raporda işaretli grup sayısı: {len(isaretli)}")

# 2) Dosya Excel'de açıksa çalışma
try:
    f = open(DOSYA, "r+b")
    f.close()
except PermissionError:
    print("HATA: '2026 DEPO STOK.xlsm' şu an Excel'de açık görünüyor.")
    print("Önce Excel'i tamamen kapatın, sonra bu programı tekrar çalıştırın.")
    sys.exit(1)

# 3) Silinecek satırları dosyanın GÜNCEL halinden hesapla + içerik doğrula
print("Dosya taranıyor (büyük dosya, biraz sürebilir)...")
wb = openpyxl.load_workbook(DOSYA, read_only=True, data_only=True)
ws = wb["STOK GİRİŞ ÇIKIŞ"]
gruplar = defaultdict(list)
for sn, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    tip, personel, stok, adet, plaka, kalan, tarih = (row + (None,) * 10)[:7]
    if not (tip and personel and stok and tarih):
        continue
    if str(tip).strip().upper() != "ÇIKIŞ":
        continue
    t = tarih.date() if hasattr(tarih, "date") else tarih
    anahtar = (str(personel).strip().upper(), str(stok).strip().upper(), t)
    if anahtar in isaretli:
        gruplar[anahtar].append(sn)
wb.close()

sil_satirlar = set()
bulunamayan = isaretli - set(gruplar)
tekli = []
for anahtar, satirlar in gruplar.items():
    if len(satirlar) > 1:
        sil_satirlar.update(sorted(satirlar)[1:])  # ilk kayıt kalır
    else:
        tekli.append(anahtar)
for a in sorted(bulunamayan):
    print(f"NOT: {a[0]} / {a[1]} / {a[2]:%d.%m.%Y} dosyada bulunamadı (belki silinmiş), atlandı.")
for a in sorted(tekli):
    print(f"NOT: {a[0]} / {a[1]} / {a[2]:%d.%m.%Y} zaten tek kayıt, silinecek kopyası yok.")

print(f"Silinecek satır sayısı: {len(sil_satirlar)}")
if not sil_satirlar:
    print("Silinecek bir şey kalmadı — dosyaya dokunulmadı.")
    sys.exit(0)

# 4) Otomatik yedek
yedek = DOSYA.replace(".xlsm", f"_YEDEK_{datetime.now():%Y-%m-%d_%H%M}.xlsm")
shutil.copy2(DOSYA, yedek)
print(f"Yedek alındı: {yedek}")

# 5) XML cerrahi düzenleme: önce geçici dosyaya yaz, sonra yerine koy
gecici = DOSYA.replace(".xlsm", "_GECICI_YENI.xlsm")
ozet = satirlari_sil(DOSYA, gecici, sil_satirlar)
print(f"Satırlar silindi: {ozet['eski_max']} -> {ozet['yeni_max']} satır.")

# 6) Sonucu doğrula (openpyxl ile açılıyor mu, satır sayısı doğru mu)
try:
    kwb = openpyxl.load_workbook(gecici, read_only=True, keep_vba=True)
    kws = kwb["STOK GİRİŞ ÇIKIŞ"]
    yeni_satir = kws.max_row
    kwb.close()
except Exception as e:
    os.remove(gecici)
    print(f"HATA: Yeni dosya doğrulanamadı ({e}). Dosyaya dokunulmadı, yedek duruyor.")
    sys.exit(1)

beklenen = ozet["eski_max"] - len(sil_satirlar)
if yeni_satir != beklenen:
    os.remove(gecici)
    print(f"HATA: Satır sayısı beklenenle uyuşmadı ({yeni_satir} != {beklenen}). İptal edildi.")
    sys.exit(1)

os.replace(gecici, DOSYA)
print()
print(f"BİTTİ: {len(sil_satirlar)} mükerrer satır silindi, dosya kaydedildi.")
print("Makrolar ve butonlar korundu. Dosyayı Excel'de açtığınızda formüller")
print("kendiliğinden yeniden hesaplanır.")
print(f"Bir sorun olursa yedekten dönebilirsiniz: {yedek}")
