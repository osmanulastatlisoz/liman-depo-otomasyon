# -*- coding: utf-8 -*-
"""2026 DEPO STOK.xlsm - STOK GİRİŞ ÇIKIŞ sayfasında mükerrer çıkış tespiti.
Aynı tarih + aynı personel + aynı stok adıyla birden fazla yazılmış ÇIKIŞ
kayıtlarını bulur. Dosyayı DEĞİŞTİRMEZ, sadece okur ve rapor üretir.
Rapordaki SİL? sütununa E yazılan gruplar mukerrer_sil.py ile silinir."""
import os
import sys
import json
import warnings
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8")

AYAR_JSON = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
                         "DepoAraclari", "yollar.json")


def _kaynak_bul():
    """2026 DEPO STOK.xlsm: kayıtlı ayar -> masaüstü -> kullanıcıya sor
    (seçim yollar.json'a kaydedilir; araç her bilgisayarda çalışır)."""
    try:
        with open(AYAR_JSON, encoding="utf-8") as f:
            kayit = json.load(f)
    except Exception:
        kayit = {}
    yol = kayit.get("depo_stok_2026")
    if yol and os.path.exists(yol):
        return yol
    varsayilan = os.path.join(os.path.expanduser("~"), "Desktop", "2026 DEPO STOK.xlsm")
    if os.path.exists(varsayilan):
        return varsayilan
    import tkinter as tk
    from tkinter import filedialog
    kok = tk.Tk(); kok.withdraw()
    yol = filedialog.askopenfilename(title="2026 DEPO STOK.xlsm dosyasını seç",
                                     filetypes=[("Excel", "*.xlsm *.xlsx")])
    kok.destroy()
    if not yol:
        sys.exit("Dosya seçilmedi, çıkılıyor.")
    kayit["depo_stok_2026"] = yol
    try:
        os.makedirs(os.path.dirname(AYAR_JSON), exist_ok=True)
        with open(AYAR_JSON, "w", encoding="utf-8") as f:
            json.dump(kayit, f, ensure_ascii=False, indent=1)
    except OSError:
        pass
    return yol


KAYNAK = _kaynak_bul()
RAPOR = os.path.join(os.path.dirname(KAYNAK), "MUKERRER_KAYIT_RAPORU.xlsx")

wb = openpyxl.load_workbook(KAYNAK, read_only=True, data_only=True)
ws = wb["STOK GİRİŞ ÇIKIŞ"]

gruplar = defaultdict(list)
for satir_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    tip, personel, stok, adet, plaka, kalan, tarih, ara, kod, aciklama = (row + (None,) * 10)[:10]
    if tip is None or personel is None or stok is None or tarih is None:
        continue
    if str(tip).strip().upper() != "ÇIKIŞ":
        continue
    t = tarih.date() if hasattr(tarih, "date") else tarih
    anahtar = (str(personel).strip().upper(), str(stok).strip().upper(), t)
    gruplar[anahtar].append((satir_no, adet, aciklama, str(personel).strip(), str(stok).strip()))
wb.close()

mukerrer = {k: v for k, v in gruplar.items() if len(v) > 1}

rw = openpyxl.Workbook()

# --- Kullanım sayfası ---
ks = rw.active
ks.title = "NASIL KULLANILIR"
adimlar = [
    "MÜKERRER KAYIT SİLME - KULLANIM ADIMLARI",
    "",
    "1) 'MÜKERRER KAYITLAR' sayfasına geçin ve listeyi inceleyin.",
    "2) Silinmesini istediğiniz grubun SİL? hücresine E yazın",
    "   (grubun tek bir satırına yazmanız yeterli).",
    "3) Bu raporu KAYDEDİN ve KAPATIN.",
    "4) '2026 DEPO STOK.xlsm' Excel'de açıksa KAPATIN.",
    "5) DEPO ARAÇLARI klasöründeki '2_MUKERRERLERI SIL' e çift tıklayın.",
    "",
    "ÖNEMLİ BİLGİLER:",
    "- İşaretlenen gruptan İLK kayıt KALIR, fazla kopyalar silinir.",
    "- Silmeden önce otomatik tarihli yedek alınır.",
    "- Eldiven, peçete gibi sarf malzemeleri aynı gün içinde gerçekten",
    "  iki kez verilmiş olabilir - bunlar hata değildir, işaretlemeyin.",
    "- '1_TARAMA YAP' tekrar çalıştırılırsa bu rapor SIFIRDAN yazılır,",
    "  işaretleriniz silinir. Önce işaretleyip silme işlemini bitirin.",
]
for i, satir in enumerate(adimlar, start=1):
    h = ks.cell(row=i, column=1, value=satir)
    h.font = Font(name="Arial", bold=(i == 1 or satir == "ÖNEMLİ BİLGİLER:"), size=12 if i == 1 else 11)
ks.column_dimensions["A"].width = 70

# --- Mükerrer listesi ---
rs = rw.create_sheet("MÜKERRER KAYITLAR")
basliklar = ["SİL?", "GRUP", "TARİH", "PERSONEL", "STOK ADI", "KATEGORİ", "ADET", "KAÇ KEZ", "FAZLA DÜŞÜLEN ADET"]
rs.append(basliklar)
for c in rs[1]:
    c.font = Font(name="Arial", bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="C00000")
    c.alignment = Alignment(horizontal="center")
sari = PatternFill("solid", start_color="FFF2CC")
mavi = PatternFill("solid", start_color="DDEBF7")

grup_no = 0
for (p_u, s_u, t), kayitlar in sorted(mukerrer.items(), key=lambda x: (x[0][2], x[0][0])):
    grup_no += 1
    adetler = [k[1] or 0 for k in kayitlar]
    fazla = sum(adetler) - adetler[0]
    for i, (satir_no, adet, aciklama, personel, stok) in enumerate(kayitlar):
        rs.append([None, grup_no, t.strftime("%d.%m.%Y"), personel, stok,
                   aciklama, adet, len(kayitlar), fazla if i == 0 else None])
    if grup_no % 2 == 1:
        for r in range(rs.max_row - len(kayitlar) + 1, rs.max_row + 1):
            for c in rs[r]:
                c.fill = sari
son_satir = rs.max_row
for r in range(2, son_satir + 1):
    h = rs.cell(row=r, column=1)
    h.fill = mavi
    h.alignment = Alignment(horizontal="center")
if son_satir > 1:
    dv = DataValidation(type="list", formula1='"E"', allow_blank=True, showErrorMessage=False)
    rs.add_data_validation(dv)
    dv.add(f"A2:A{son_satir}")
for c_idx, w in zip("ABCDEFGHI", [6, 7, 12, 26, 36, 26, 7, 9, 18]):
    rs.column_dimensions[c_idx].width = w
for row in rs.iter_rows(min_row=2):
    for c in row:
        if c.font.bold is not True:
            c.font = Font(name="Arial")
rs.freeze_panes = "A2"

rw.save(RAPOR)

print(f"Aynı gün + aynı personel + aynı stok tekrarı: {len(mukerrer)} grup bulundu.")
print()
print("Şimdi açılacak raporda silinmesini istediklerinizin SİL? hücresine E yazın,")
print("raporu kaydedip kapatın, sonra '2_MUKERRERLERI SIL' i çalıştırın.")
print("(Eldiven/peçete gibi sarf malzemeleri aynı gün iki kez verilmiş olabilir —")
print(" emin olmadıklarınızı işaretlemeyin.)")
print()
print(f"Rapor kaydedildi: {RAPOR}")
