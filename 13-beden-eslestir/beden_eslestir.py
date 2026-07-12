# -*- coding: utf-8 -*-
"""
KIŞLIK BEDEN LİSTESİ — isim düzeltme + ayrılmışları çıkarma aracı
(dolap_eslestir.py'nin beden listesi uyarlaması)

Girdiler (bu script'in bulunduğu klasörde):
  - KIŞLIK BEDEN LİSTESİ.xlsx  -> "KIŞLIK BEDEN LİSTESİ" sayfası (A=isim, B-F beden/renk)
  - DOĞRU PERSONEL İSİMLERİ.xlsx  -> güncel personel (AD-SOYAD, TC)
  - AYRILMIŞ PERSONEL LİSTESİ.xlsx -> işten ayrılanlar (AD-SOYAD, TC)

İşleyiş:
  1) Güncel listeyle birebir eşleşen -> KALIR (yazım farkı varsa düzeltilir).
     Ayrılmış listeyle birebir eşleşen -> ÇIKARILIR (otomatik).
     İki listede birden olan güncel sayılır.
  2) Kalan isimler SIRAYLA sorulur: öneriler güncel + [AYRILMIŞ] etiketli;
     ↓ seç + Enter = eşleştir; yazınca canlı arama; "Hiçbir listede yok" =
     güncel değil -> listeden çıkarılır.
  3) Kararlar beden_kararlar.json'a anında yazılır (yarıda bırak, devam et).
  4) Bitince: yedek alınır, düzeltmeler ve satır silmeleri AÇIK Excel'de
     UYGULANIR (kaydetmez — kontrol edip Ctrl+S sana kalır) ve
     BEDEN_DUZELTME_SONUC.xlsx raporu üretilir (çıkarılanların beden
     bilgileri raporda saklanır, veri kaybolmaz).
"""
import os
import sys
import json
import time
import shutil
import difflib
import tempfile
import unicodedata
import warnings
import datetime

warnings.filterwarnings("ignore")

BURA = os.path.dirname(os.path.abspath(__file__))

# ======================= AYARLAR =======================
# Başka bir listeye (örn. YAZLIK) uyarlamak için genelde şu üçünü
# değiştirmek yeter. Tüm dosyalar script'in bulunduğu klasörde aranır.
BEDEN = os.path.join(BURA, "KIŞLIK BEDEN LİSTESİ.xlsx")   # düzeltilecek dosya
SAYFA = "KIŞLIK BEDEN LİSTESİ"                            # isimlerin olduğu sayfa
VERI_SUTUN = 6   # A..F: A=isim + rapora yedeklenecek veri sütunu sayısı

DOGRU_X = os.path.join(BURA, "DOĞRU PERSONEL İSİMLERİ.xlsx")
AYRILMIS_X = os.path.join(BURA, "AYRILMIŞ PERSONEL LİSTESİ.xlsx")
KARAR_JSON = os.path.join(BURA, "beden_kararlar.json")
RAPOR_XLSX = os.path.join(BURA, "BEDEN_DUZELTME_SONUC.xlsx")
# ========================================================

_TR = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")


def fold(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return " ".join(s.translate(_TR).upper().split())


def _kopyadan_ac(yol):
    from openpyxl import load_workbook
    kopya = os.path.join(tempfile.gettempdir(), f"beden_{os.getpid()}_{os.path.basename(yol)}")
    shutil.copy2(yol, kopya)
    return load_workbook(kopya, read_only=True, data_only=True), kopya


def veri_yukle():
    dogru, tc = {}, {}
    wb, k = _kopyadan_ac(DOGRU_X)
    for r in wb.active.iter_rows(min_row=2, values_only=True):
        if r[0] and str(r[0]).strip():
            f = fold(r[0])
            dogru.setdefault(f, str(r[0]).strip())
            if len(r) > 1 and r[1]:
                tc.setdefault(f, r[1])
    wb.close(); os.remove(k)

    ayrilmis, ayr_tc = {}, {}
    wb, k = _kopyadan_ac(AYRILMIS_X)
    for r in wb.active.iter_rows(min_row=2, values_only=True):
        if r[0] and str(r[0]).strip():
            f = fold(r[0])
            if f in dogru:      # iki listede birden -> güncel sayılır
                continue
            ayrilmis.setdefault(f, str(r[0]).strip())
            if len(r) > 1 and r[1]:
                ayr_tc.setdefault(f, r[1])
    wb.close(); os.remove(k)

    beden = []   # {satir, ad, veri: (A..F)}
    wb, k = _kopyadan_ac(BEDEN)
    ws = wb[SAYFA]
    for i, r in enumerate(ws.iter_rows(min_row=2, max_col=VERI_SUTUN, values_only=True), start=2):
        if r[0] and str(r[0]).strip():
            beden.append({"satir": i, "ad": str(r[0]).strip(), "veri": r})
    wb.close(); os.remove(k)
    return dogru, tc, ayrilmis, ayr_tc, beden


def kararlari_oku():
    try:
        with open(KARAR_JSON, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def kararlari_yaz(kararlar):
    gecici = KARAR_JSON + ".tmp"
    with open(gecici, "w", encoding="utf-8") as f:
        json.dump(kararlar, f, ensure_ascii=False, indent=1)
    os.replace(gecici, KARAR_JSON)


def anahtar(kayit):
    return f"{kayit['satir']}|{fold(kayit['ad'])}"


def siniflandir(dogru, ayrilmis, beden, kararlar):
    """Döner: duzeltmeler [(satır, eski, yeni, tür)],
    cikarilacaklar [(satır, dolaptaki ad, listedeki ad, tür, veri)], kararsiz."""
    duzeltmeler = []
    cikarilacaklar = []
    kararsiz = []
    for d in beden:
        f = fold(d["ad"])
        kr = kararlar.get(anahtar(d))
        if f in dogru:
            if d["ad"] != dogru[f]:
                duzeltmeler.append((d["satir"], d["ad"], dogru[f], "OTOMATİK (yazım)"))
        elif kr and kr["karar"] == "ESLE" and kr["hedef"] in dogru:
            duzeltmeler.append((d["satir"], d["ad"], dogru[kr["hedef"]], "ELLE EŞLEŞTİRME"))
        elif kr and kr["karar"] == "AYRILDI" and kr["hedef"] in ayrilmis:
            cikarilacaklar.append((d["satir"], d["ad"], ayrilmis[kr["hedef"]],
                                   "ELLE (ayrılmış listesi)", d["veri"]))
        elif f in ayrilmis:
            cikarilacaklar.append((d["satir"], d["ad"], ayrilmis[f],
                                   "OTOMATİK (ayrılmış listesi)", d["veri"]))
        elif kr and kr["karar"] == "YOK":
            cikarilacaklar.append((d["satir"], d["ad"], "", "HİÇBİR LİSTEDE YOK", d["veri"]))
        else:
            kararsiz.append(d)
    return duzeltmeler, cikarilacaklar, kararsiz


# ------------------------------------------------------------------ rapor
def rapor_yaz(dogru, tc, ayrilmis, ayr_tc, beden, kararlar):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    duzeltmeler, cikarilacaklar, kararsiz = siniflandir(dogru, ayrilmis, beden, kararlar)

    wb = Workbook()

    def sayfa(ad, basliklar, satirlar, genislikler):
        ws = wb.create_sheet(ad) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = ad
        ws.append(basliklar)
        for c in ws[1]:
            c.font = Font(name="Arial", bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="1F4E78")
        for r in satirlar:
            ws.append(r)
        for i, g in enumerate(genislikler, 1):
            ws.column_dimensions[get_column_letter(i)].width = g
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(basliklar))}{ws.max_row}"

    sayfa("DÜZELTİLENLER",
          ["SATIR", "ESKİ İSİM", "DOĞRU İSİM", "TÜR"],
          duzeltmeler, [8, 30, 30, 24])
    sayfa("ÇIKARILANLAR",
          ["SATIR", "LİSTEDEKİ İSİM", "DOĞRU YAZIMI", "TC",
           "PANTOLON", "KAZAK", "MONT", "TULUM", "RENK", "TÜR"],
          [(s, a, la, ayr_tc.get(fold(la), ""),
            v[1], v[2], v[3], v[4], v[5], t)
           for s, a, la, t, v in cikarilacaklar],
          [8, 30, 30, 15, 11, 9, 9, 9, 12, 26])
    if kararsiz:
        sayfa("KARARSIZ", ["SATIR", "İSİM"],
              [(d["satir"], d["ad"]) for d in kararsiz], [8, 32])
    sayfa("ÖZET", ["BİLGİ", "SAYI"], [
        ("Beden listesi kayıt", len(beden)),
        ("Kalan (güncel personel)", len(beden) - len(cikarilacaklar) - len(kararsiz)),
        ("Düzeltilen isim", len(duzeltmeler)),
        ("Çıkarılan (ayrılmış/listede yok)", len(cikarilacaklar)),
        ("Kararsız", len(kararsiz)),
        ("Rapor tarihi", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")),
    ], [34, 18])

    gecici = RAPOR_XLSX + ".tmp"
    wb.save(gecici)
    os.replace(gecici, RAPOR_XLSX)
    return len(duzeltmeler), len(cikarilacaklar), len(kararsiz)


# ------------------------------------------------------------------ uygula
def uygula(dogru, ayrilmis, beden, kararlar, kaydet=False):
    """Yedek alır; isimleri düzeltir, çıkarılacak satırları siler (COM,
    biçimler korunur). kaydet=False: dosya AÇIK ve KAYDEDİLMEMİŞ bırakılır —
    kontrol edip Ctrl+S sana kalır."""
    import threading
    import win32com.client
    import win32gui, win32process, win32con, pywintypes

    duzeltmeler, cikarilacaklar, _ = siniflandir(dogru, ayrilmis, beden, kararlar)
    if not duzeltmeler and not cikarilacaklar:
        return 0, 0, None

    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
    yedek = os.path.join(BURA, f"KIŞLIK BEDEN LİSTESİ_YEDEK_{ts}.xlsx")
    if not os.path.exists(yedek):
        shutil.copy2(BEDEN, yedek)

    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        excel.Workbooks.Count          # sağlık kontrolü (zombi örnek olabilir)
        yeni = False
    except Exception:
        excel = win32com.client.DispatchEx("Excel.Application")
        yeni = True
        time.sleep(2)

    # etkinleştirme sihirbazı gözcüsü (yeni örnek açılırsa gelebilir)
    surdur = [True]
    def izleyici():
        def kapat(h, _):
            try:
                _, p = win32process.GetWindowThreadProcessId(h)
                if win32gui.GetClassName(h) == "NUIDialog":
                    win32gui.PostMessage(h, win32con.WM_CLOSE, 0, 0)
            except Exception:
                pass
        while surdur[0]:
            try:
                win32gui.EnumWindows(kapat, None)
            except Exception:
                pass
            time.sleep(0.4)
    threading.Thread(target=izleyici, daemon=True).start()

    def dene(f, saniye=40):
        son = None
        for _ in range(saniye * 2):
            try:
                return f()
            except pywintypes.com_error as e:
                son = e
                time.sleep(0.5)
        raise son

    try:
        try:
            excel.Visible = True       # kozmetik; zombi örnekte başarısız olabilir
        except Exception:
            pass
        hedef_ad = fold(os.path.basename(BEDEN))
        wb = None
        for w in excel.Workbooks:
            if fold(w.Name) == hedef_ad:
                wb = w
                break
        if wb is None:
            wb = dene(lambda: excel.Workbooks.Open(BEDEN, 0))
        ws = wb.Worksheets(SAYFA)

        d_n = 0
        for satir, eski, yenia, tur in duzeltmeler:
            hucre = ws.Cells(satir, 1)
            if hucre.Value is not None and fold(hucre.Value) == fold(eski):
                hucre.Value = yenia
                d_n += 1

        s_n = 0
        for satir, ad, la, tur, veri in sorted(cikarilacaklar, reverse=True):
            hucre = ws.Cells(satir, 1)
            if hucre.Value is not None and fold(hucre.Value) in (fold(ad), fold(la)):
                ws.Rows(satir).Delete()
                s_n += 1

        if kaydet:
            dene(lambda: wb.Save())
            dene(lambda: wb.Close(False))
            if yeni:
                dene(lambda: excel.Quit(), 10)
        return d_n, s_n, None
    except Exception as e:
        return None, None, str(e)
    finally:
        surdur[0] = False


# ------------------------------------------------------------------ GUI
def gui():
    import tkinter as tk
    from tkinter import ttk, messagebox

    for yol in (BEDEN, DOGRU_X, AYRILMIS_X):
        if not os.path.exists(yol):
            r = tk.Tk(); r.withdraw()
            messagebox.showerror("Dosya yok", f"Bulunamadı:\n{yol}")
            return

    dogru, tc, ayrilmis, ayr_tc, beden = veri_yukle()
    kararlar = kararlari_oku()
    havuz = {k: (v, "AYRILMIŞ") for k, v in ayrilmis.items()}
    havuz.update({k: (v, "GÜNCEL") for k, v in dogru.items()})
    anahtar_liste = list(havuz)

    tam = sum(1 for d in beden if fold(d["ad"]) in dogru)
    oto_ayr = sum(1 for d in beden
                  if fold(d["ad"]) not in dogru and fold(d["ad"]) in ayrilmis
                  and anahtar(d) not in kararlar)
    kuyruk = [d for d in beden
              if fold(d["ad"]) not in dogru and fold(d["ad"]) not in ayrilmis
              and anahtar(d) not in kararlar]

    root = tk.Tk()
    root.title("KIŞLIK BEDEN LİSTESİ — İSİM DÜZELTME")
    root.geometry("780x640")
    stil = ttk.Style(root)
    try:
        stil.theme_use("vista")
    except Exception:
        pass

    durum = {"i": 0, "gecmis": []}

    ust_var = tk.StringVar()
    ttk.Label(root, textvariable=ust_var, font=("Segoe UI", 11),
              padding=(14, 8, 14, 0)).pack(fill="x")
    ad_var = tk.StringVar()
    ttk.Label(root, textvariable=ad_var, font=("Segoe UI", 20, "bold"),
              foreground="#B33", padding=(14, 2)).pack(fill="x")
    bilgi_var = tk.StringVar()
    ttk.Label(root, textvariable=bilgi_var, font=("Segoe UI", 10),
              foreground="#555", padding=(14, 0)).pack(fill="x")

    ara_cerceve = ttk.Frame(root, padding=(14, 8, 14, 2))
    ara_cerceve.pack(fill="x")
    ttk.Label(ara_cerceve, text="Ara:", font=("Segoe UI", 12)).pack(side="left")
    ara_var = tk.StringVar()
    ara = ttk.Entry(ara_cerceve, textvariable=ara_var, font=("Segoe UI", 14))
    ara.pack(side="left", fill="x", expand=True, padx=(8, 0))

    kutu = tk.Listbox(root, font=("Segoe UI", 13), height=10,
                      selectbackground="#1F4E78", selectforeground="white")
    kutu.pack(fill="both", expand=True, padx=14, pady=(4, 2))
    gercek = []

    ipucu = ("↓ ↑ listeden seç   Enter = EŞLEŞTİR   |   kutuya yazınca canlı arama\n"
             "[AYRILMIŞ] etiketli seçim = listeden çıkarılır   |   "
             "sağdaki düğme: hiçbir listede yok (çıkarılır)")
    ttk.Label(root, text=ipucu, font=("Segoe UI", 10), foreground="#555",
              padding=(14, 2)).pack(fill="x")

    dugme = ttk.Frame(root, padding=(14, 4, 14, 10))
    dugme.pack(fill="x")
    esle_btn = tk.Button(dugme, text="✔ EŞLEŞTİR (Enter)", bg="#2e7d32", fg="white",
                         font=("Segoe UI", 12, "bold"), padx=14)
    esle_btn.pack(side="left")
    yok_btn = tk.Button(dugme, text="✖ Hiçbir listede yok", bg="#b33", fg="white",
                        font=("Segoe UI", 12, "bold"), padx=14)
    yok_btn.pack(side="left", padx=10)
    geri_btn = tk.Button(dugme, text="↶ Geri", font=("Segoe UI", 11), padx=10)
    geri_btn.pack(side="left")

    alt_var = tk.StringVar()
    ttk.Label(root, textvariable=alt_var, relief="sunken", anchor="w",
              padding=(10, 4), font=("Segoe UI", 10)).pack(fill="x", side="bottom")

    def oneri_doldur(kayit):
        kutu.delete(0, "end")
        gercek.clear()
        aranan = fold(ara_var.get())
        if aranan:
            kelimeler = aranan.split()
            adaylar = [k for k in anahtar_liste if all(x in k for x in kelimeler)][:40]
        else:
            f = fold(kayit["ad"])
            adaylar = difflib.get_close_matches(f, anahtar_liste, n=8, cutoff=0.5)
            kel = set(f.split())
            for k in anahtar_liste:
                if k not in adaylar and kel & set(k.split()):
                    adaylar.append(k)
            adaylar = adaylar[:15]
        f0 = fold(kayit["ad"])
        oranlar = []
        for k in adaylar:
            gercek.append(k)
            ad, kaynak = havuz[k]
            oran = int(difflib.SequenceMatcher(None, f0, k).ratio() * 100)
            oranlar.append(oran)
            etiket = "  [AYRILMIŞ]" if kaynak == "AYRILMIŞ" else ""
            kutu.insert("end", f"{ad}{etiket}    (%{oran} benzer)")
        if kutu.size() and not aranan and oranlar[0] >= 80:
            kutu.selection_set(0)

    def goster():
        if durum["i"] >= len(kuyruk):
            bitir()
            return
        kayit = kuyruk[durum["i"]]
        v = kayit["veri"]
        ust_var.set(f"Karar {durum['i'] + 1} / {len(kuyruk)}   "
                    f"(otomatik: {tam} güncel + {oto_ayr} ayrılmış)")
        ad_var.set(kayit["ad"])
        bilgi_var.set(f"Satır {kayit['satir']}  •  Pantolon {v[1]}  Kazak {v[2]}  "
                      f"Mont {v[3]}  Renk {v[5]}")
        ara_var.set("")
        oneri_doldur(kayit)
        ara.focus_set()

    def karar_ver(tip):
        kayit = kuyruk[durum["i"]]
        if tip == "ESLE":
            sec = kutu.curselection()
            if not sec:
                alt_var.set("Önce listeden bir isim seç (↓ tuşu) ya da arayıp seç.")
                return
            hedef = gercek[sec[0]]
            ad, kaynak = havuz[hedef]
            if kaynak == "AYRILMIŞ":
                kararlar[anahtar(kayit)] = {"karar": "AYRILDI", "hedef": hedef, "hedef_ad": ad}
                alt_var.set(f"{kayit['ad']}  →  {ad}  [AYRILMIŞ → çıkarılacak]")
            else:
                kararlar[anahtar(kayit)] = {"karar": "ESLE", "hedef": hedef, "hedef_ad": ad}
                alt_var.set(f"{kayit['ad']}  →  {ad}")
        else:
            kararlar[anahtar(kayit)] = {"karar": "YOK"}
            alt_var.set(f"{kayit['ad']}  →  hiçbir listede yok (çıkarılacak)")
        kararlari_yaz(kararlar)
        durum["gecmis"].append(durum["i"])
        durum["i"] += 1
        goster()

    def geri(*_):
        if not durum["gecmis"]:
            return
        durum["i"] = durum["gecmis"].pop()
        kararlar.pop(anahtar(kuyruk[durum["i"]]), None)
        kararlari_yaz(kararlar)
        goster()

    def kaydir(yon):
        if not kutu.size():
            return "break"
        sec = kutu.curselection()
        i = (sec[0] + yon) if sec else 0
        i = max(0, min(kutu.size() - 1, i))
        kutu.selection_clear(0, "end")
        kutu.selection_set(i)
        kutu.see(i)
        return "break"

    def bitir():
        for w in root.winfo_children():
            w.destroy()
        ttk.Label(root, text="Eşleştirme tamamlandı ✔", font=("Segoe UI", 18, "bold"),
                  padding=(16, 16)).pack()
        uygulama_var = tk.StringVar()
        ttk.Label(root, textvariable=uygulama_var, font=("Segoe UI", 11),
                  foreground="#2e7d32", padding=(16, 2), wraplength=700).pack()
        ozet_var = tk.StringVar()
        ttk.Label(root, textvariable=ozet_var, font=("Segoe UI", 12),
                  padding=(16, 4)).pack()

        def raporla():
            dz, ck, ks = rapor_yaz(dogru, tc, ayrilmis, ayr_tc, beden, kararlar)
            ozet_var.set(f"Düzeltilen isim: {dz}   |   çıkarılan: {ck}   |   kararsız: {ks}\n"
                         f"Rapor: {os.path.basename(RAPOR_XLSX)}")
            os.startfile(RAPOR_XLSX)

        def uygula_tikla(sessiz=False):
            d_n, s_n, hata = uygula(dogru, ayrilmis, beden, kararlar)
            if hata:
                uygulama_var.set("⚠ Uygulanamadı: " + hata)
                if not sessiz:
                    messagebox.showerror("Uygulanamadı", hata)
            else:
                uygulama_var.set(f"✏ {d_n} isim düzeltildi, {s_n} satır çıkarıldı — "
                                 "dosya AÇIK ve KAYDEDİLMEDİ; kontrol edip Ctrl+S ile kaydet. "
                                 "(Yedek aynı klasörde)")

        tk.Button(root, text="📄 Rapor oluştur ve aç", command=raporla, bg="#1F4E78",
                  fg="white", font=("Segoe UI", 13, "bold"), padx=16, pady=6).pack(pady=8)
        tk.Button(root, text="✏ Düzeltme + çıkarmayı yeniden uygula",
                  command=uygula_tikla, font=("Segoe UI", 11), padx=12).pack(pady=4)
        ttk.Label(root, text="Kararları değiştirmek için beden_kararlar.json'u silip aracı"
                             " yeniden aç. Uygulamayı BİR kez yap (satır numaraları kayar).",
                  font=("Segoe UI", 9), foreground="#777", padding=(10, 10)).pack()
        root.after(300, lambda: uygula_tikla(sessiz=True))

    esle_btn.config(command=lambda: karar_ver("ESLE"))
    yok_btn.config(command=lambda: karar_ver("YOK"))
    geri_btn.config(command=geri)
    ara.bind("<Return>", lambda e: karar_ver("ESLE"))
    ara.bind("<Down>", lambda e: kaydir(1))
    ara.bind("<Up>", lambda e: kaydir(-1))
    ara.bind("<KeyRelease>", lambda e: None if e.keysym in
             ("Return", "Up", "Down", "Escape") else
             oneri_doldur(kuyruk[durum["i"]]) if durum["i"] < len(kuyruk) else None)
    ara.bind("<Escape>", lambda e: (ara_var.set(""),
             oneri_doldur(kuyruk[durum["i"]]) if durum["i"] < len(kuyruk) else None))
    kutu.bind("<Double-Button-1>", lambda e: karar_ver("ESLE"))

    goster()
    root.mainloop()


if __name__ == "__main__":
    gui()
