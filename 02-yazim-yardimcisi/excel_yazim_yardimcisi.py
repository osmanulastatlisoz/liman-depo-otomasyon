# -*- coding: utf-8 -*-
"""
EXCEL YAZIM YARDIMCISI
Excel'in yanında duran, hep üstte küçük bir pencere. Aktif hücre hangi
sütundaysa ona göre öneri listesi gösterir (PERSONEL, MALZEME, İŞLEM, DEPO):
  - yazdıkça liste süzülür (kelime sırası önemsiz)
  - ↑ ↓  : listeden seç
  - →    : seçileni aktif hücreye yaz ve SAĞDAKİ hücreye geç
  - ←    : yaz ve SOLDAKİ hücreye geç
  - Enter: yaz ve sağa geç
  - Esc  : kutuyu temizle
  - kutu BOŞKEN → / ← : yazmadan sadece hücre değiştir
Excel tüm özellikleriyle açık kalır; pencere sadece aktif hücreye değer yazar.

Kullanım: YAZIM YARDIMCISI.bat (çift tık) veya python excel_yazim_yardimcisi.py
"""
import os
import re
import json
import shutil
import tempfile
import unicodedata
import warnings
import datetime
from collections import Counter

warnings.filterwarnings("ignore")

MASAUSTU = os.path.join(os.path.expanduser("~"), "Desktop")
AYAR_JSON = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
                         "DepoAraclari", "yollar.json")


def kayitli_yol(anahtar, varsayilanlar, baslik):
    """Dosyayı sırayla arar: kayıtlı ayar -> varsayılan konumlar -> kullanıcıya
    sor. Seçim yollar.json'a kaydedilir, bir daha sorulmaz — böylece araç
    farklı klasör düzenine sahip bilgisayarlarda da çalışır."""
    try:
        with open(AYAR_JSON, encoding="utf-8") as f:
            kayit = json.load(f)
    except Exception:
        kayit = {}
    yol = kayit.get(anahtar)
    if yol and os.path.exists(yol):
        return yol
    for v in varsayilanlar:
        if os.path.exists(v):
            return v
    import tkinter as tk
    from tkinter import filedialog
    kok = tk.Tk(); kok.withdraw()
    yol = filedialog.askopenfilename(
        title=baslik, filetypes=[("Excel", "*.xlsm *.xlsx"), ("Tümü", "*.*")])
    kok.destroy()
    if not yol:
        raise SystemExit(f"{baslik} — dosya seçilmedi, çıkılıyor.")
    kayit[anahtar] = yol
    try:
        os.makedirs(os.path.dirname(AYAR_JSON), exist_ok=True)
        with open(AYAR_JSON, "w", encoding="utf-8") as f:
            json.dump(kayit, f, ensure_ascii=False, indent=1)
    except OSError:
        pass
    return yol


DOSYA_TCH = kayitli_yol("tch_stok_takip",
                        [os.path.join(MASAUSTU, "TCH_STOK_TAKIP_CALISMASI GÜNCEL.xlsm")],
                        "TCH_STOK_TAKIP_CALISMASI GÜNCEL.xlsm dosyasını seç")

# sayfa adı -> {sütun no: rol}
SAYFA_KOLON = {
    "STOK HAREKETLERİ": {1: "TARİH", 2: "HEDEF DEPO", 3: "İŞLEM TÜRÜ",
                         4: "PERSONEL", 5: "MALZEME", 6: "MİKTAR", 7: "AÇIKLAMA / PLAKA"},
    "STOK GİRİŞ ÇIKIŞ": {1: "İŞLEM TÜRÜ", 2: "PERSONEL", 3: "MALZEME",
                         4: "MİKTAR", 5: "AÇIKLAMA / PLAKA", 7: "TARİH"},
}
LISTE_ROLLER = ("PERSONEL", "MALZEME", "İŞLEM TÜRÜ", "HEDEF DEPO")

_TR = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")


def fold(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return " ".join(s.translate(_TR).upper().split())


def listeleri_oku():
    """Öneri listelerini TCH dosyasının resmi liste sayfalarından okur
    (dosya Excel'de açıkken de çalışsın diye geçici kopyadan)."""
    from openpyxl import load_workbook
    kopya = os.path.join(tempfile.gettempdir(), f"eyy_{os.getpid()}_listeler.xlsm")
    shutil.copy2(DOSYA_TCH, kopya)
    try:
        wb = load_workbook(kopya, read_only=True, data_only=True)

        def sutun(sayfa, kolon, min_row=2):
            gorulen, sira = set(), []
            for r in wb[sayfa].iter_rows(min_row=min_row, min_col=kolon,
                                         max_col=kolon, values_only=True):
                v = str(r[0]).strip() if r[0] is not None else ""
                if v and fold(v) not in gorulen:
                    gorulen.add(fold(v))
                    sira.append(v)
            return sira

        listeler = {
            "PERSONEL": sorted(sutun("PERSONEL LİSTESİ", 2), key=fold),
            "MALZEME": sorted(sutun("MALZEME LİSTESİ", 1), key=fold),
            "HEDEF DEPO": sutun("LISTELER", 1),
            "İŞLEM TÜRÜ": sutun("LISTELER", 2),
        }
        # DEPO sayfası: A=MALZEME ADI, G=LİMAN DEPO (kalan miktar)
        depo = {}
        for r in wb["DEPO"].iter_rows(min_row=2, min_col=1, max_col=7, values_only=True):
            if r[0]:
                depo[fold(r[0])] = r[6]
        listeler["MALZEME_STOK"] = depo
        # STOK HAREKETLERİ: D=PERSONEL — kişi başına hareket sayısı (sıralama için)
        sayilar = Counter()
        for r in wb["STOK HAREKETLERİ"].iter_rows(min_row=2, min_col=4, max_col=4,
                                                  values_only=True):
            if r[0]:
                sayilar[fold(r[0])] += 1
        listeler["PERSONEL_SAYI"] = dict(sayilar)
        wb.close()
        return listeler
    finally:
        try:
            os.remove(kopya)
        except OSError:
            pass


def stok_metni(v):
    """Stok sayısını kısa metne çevirir: 4 -> '4', 2.5 -> '2,5'."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v)
    return str(int(f)) if f == int(f) else f"{f:.1f}".replace(".", ",")


def depo_stok_com(excel):
    """Excel'de AÇIK duran TCH dosyasının DEPO sayfasından liman stoğunu
    CANLI okur (tek toplu COM çağrısı). Olmazsa None döner."""
    try:
        hedef_ad = fold(os.path.basename(DOSYA_TCH))
        wb = None
        for w in excel.Workbooks:
            if fold(w.Name) == hedef_ad:
                wb = w
                break
        if wb is None:
            return None
        ws = wb.Worksheets("DEPO")
        son = ws.Cells(ws.Rows.Count, 1).End(-4162).Row
        if son < 3:
            return None
        veri = ws.Range(ws.Cells(2, 1), ws.Cells(son, 7)).Value
        return {fold(r[0]): r[6] for r in veri if r and r[0]}
    except Exception:
        return None


def tarih_cevir(metin):
    metin = metin.strip()
    for kalip in ("%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(metin, kalip)
        except ValueError:
            pass
    m = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})", metin)
    if m:
        try:
            return datetime.datetime(datetime.date.today().year, int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


def hucre_bilgi(excel):
    """(sayfa adı, adres, satır, sütun, rol) — COM hatasında None döner."""
    hucre = excel.ActiveCell
    if hucre is None:
        return None
    sayfa = hucre.Worksheet.Name
    satir, kolon = hucre.Row, hucre.Column
    from openpyxl.utils import get_column_letter
    adres = f"{get_column_letter(kolon)}{satir}"
    rol = SAYFA_KOLON.get(sayfa, {}).get(kolon, "SERBEST")
    return sayfa, adres, satir, kolon, rol


def hucreye_yaz(excel, metin, dx, dy):
    """Metni aktif hücreye yazar (boşsa yazmaz) ve seçimi (dx, dy) kadar taşır."""
    hucre = excel.ActiveCell
    ws = hucre.Worksheet
    satir, kolon = hucre.Row, hucre.Column
    metin = metin.strip()
    if metin:
        rol = SAYFA_KOLON.get(ws.Name, {}).get(kolon, "SERBEST")
        if rol == "TARİH":
            t = tarih_cevir(metin)
            hucre.Value = t if t else metin
        elif re.fullmatch(r"\d+", metin):
            hucre.Value = int(metin)
        elif re.fullmatch(r"\d+[.,]\d+", metin):
            hucre.Value = float(metin.replace(",", "."))
        else:
            hucre.Value = metin
    ws.Cells(max(1, satir + dy), max(1, kolon + dx)).Select()


def tr_upper(s):
    """Türkçe-doğru büyük harf: i→İ, ı→I; fazla boşlukları toplar."""
    return " ".join(str(s).replace("i", "İ").replace("ı", "I").upper().split())


def personel_listeye_ekle(excel, ad):
    """Adı AÇIK TCH dosyasının PERSONEL LİSTESİ sayfasında B sütununun
    sonuna ekler. (Önce liste, sonra hücre — koruma makrosu itiraz etmesin.)"""
    hedef = fold(os.path.basename(DOSYA_TCH))
    wb = None
    for w in excel.Workbooks:
        if fold(w.Name) == hedef:
            wb = w
            break
    if wb is None:
        return None, "TCH dosyası açık Excel'de bulunamadı."
    ws = wb.Worksheets("PERSONEL LİSTESİ")
    son = ws.Cells(ws.Rows.Count, 2).End(-4162).Row   # xlUp
    ws.Cells(son + 1, 2).Value = ad
    return son + 1, None


def ustu_kopyala(excel):
    """Excel'in Ctrl+D'si: üstteki hücreyi (değer/formül/biçimiyle) aktif
    hücreye doldurur. En üst satırdaysa False döner."""
    h = excel.ActiveCell
    if h.Row <= 1:
        return False
    ws = h.Worksheet
    ws.Range(ws.Cells(h.Row - 1, h.Column), h).FillDown()
    return True


def excel_bul():
    import win32com.client
    try:
        return win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return None


# ------------------------------------------------------------------ GUI
def gui():
    import tkinter as tk
    from tkinter import ttk, messagebox

    try:
        listeler = listeleri_oku()
    except Exception as e:
        root = tk.Tk(); root.withdraw()
        messagebox.showerror("Hata", f"Liste sayfaları okunamadı:\n{e}")
        return

    root = tk.Tk()
    root.title("YAZIM YARDIMCISI")
    root.geometry("430x470+40+120")
    root.attributes("-topmost", True)
    root.minsize(360, 380)

    stil = ttk.Style(root)
    try:
        stil.theme_use("vista")
    except Exception:
        pass

    durum = {"excel": None, "rol": None, "sayfa": None, "adres": ""}

    baslik_var = tk.StringVar(value="Excel aranıyor…")
    ttk.Label(root, textvariable=baslik_var, font=("Segoe UI", 14, "bold"),
              padding=(12, 8, 12, 2)).pack(fill="x")

    giris_cerceve = ttk.Frame(root, padding=(12, 2))
    giris_cerceve.pack(fill="x")
    metin_var = tk.StringVar()
    giris = ttk.Entry(giris_cerceve, textvariable=metin_var, font=("Segoe UI", 15))
    giris.pack(side="left", fill="x", expand=True)
    giris.focus_set()

    def listeleri_yenile():
        try:
            listeler.update(listeleri_oku())
            oner()
            durum_var.set("Listeler yenilendi.")
        except Exception as e:
            durum_var.set(f"Liste yenilenemedi: {e}")

    ttk.Button(giris_cerceve, text="⟳", width=3, command=listeleri_yenile).pack(
        side="left", padx=(6, 0))
    yeni_btn = ttk.Button(giris_cerceve, text="➕", width=3)
    yeni_btn.pack(side="left", padx=(4, 0))

    kutu = tk.Listbox(root, font=("Segoe UI", 13), activestyle="dotbox",
                      selectbackground="#1F4E78", selectforeground="white", height=9)
    kutu.pack(fill="both", expand=True, padx=12, pady=(4, 2))

    ipucu = ("Listeden seçmezsen yazdığın AYNEN yazılır — seçmek için ↓\n"
             "→ yaz + sağa   ← yaz + sola   Enter yaz + sağa   Esc temizle\n"
             "Ctrl+D: üsttekini kopyala   |   Ctrl+N / ➕: YENİ personel "
             "(önce listeye ekler, sonra yazar)")
    ttk.Label(root, text=ipucu, font=("Segoe UI", 10), foreground="#555",
              padding=(12, 2)).pack(fill="x")

    durum_var = tk.StringVar(value="")
    ttk.Label(root, textvariable=durum_var, relief="sunken", anchor="w",
              padding=(10, 4), font=("Segoe UI", 10)).pack(fill="x", side="bottom")

    gercek = []   # listbox sırası -> hücreye yazılacak GERÇEK değer (stok eki yok)

    def oner(*_):
        kutu.delete(0, "end")
        gercek.clear()
        rol = durum["rol"]
        if rol not in LISTE_ROLLER:
            return
        adaylar = listeler.get(rol, [])
        aranan = fold(metin_var.get())
        kelimeler = aranan.split()
        stoklar = listeler.get("MALZEME_STOK", {}) if rol == "MALZEME" else {}

        def stok_sirasi(a):
            try:
                return -float(stoklar.get(fold(a)))
            except (TypeError, ValueError):
                return float("inf")   # stoğu bilinmeyen en alta

        sayilar = listeler.get("PERSONEL_SAYI", {}) if rol == "PERSONEL" else {}
        if kelimeler:
            eslesen = [a for a in adaylar if all(k in fold(a) for k in kelimeler)]
            if rol == "MALZEME":   # stoğu çok olan üstte
                eslesen.sort(key=lambda a: (stok_sirasi(a), fold(a)))
            elif rol == "PERSONEL":   # hareketi çok olan üstte
                eslesen.sort(key=lambda a: (-sayilar.get(fold(a), 0), fold(a)))
            else:
                eslesen.sort(key=lambda a: (not fold(a).startswith(kelimeler[0]), fold(a)))
        else:
            # kısa listelerde (işlem/depo) hepsini göster, uzunlarda bekle
            eslesen = adaylar if len(adaylar) <= 12 else []
        for a in eslesen[:60]:
            gercek.append(a)
            k = fold(a)
            kutu.insert("end", f"{a}    ({stok_metni(stoklar[k])})" if k in stoklar else a)
        # bilerek ↑↓ ile seçilmedikçe hiçbir öneri seçili gelmez:
        # → ← Enter yazılan metni AYNEN geçirir (kısayol makroları bozulmasın)
        if kutu.size():
            kutu.see(0)

    def secili_metin():
        sec = kutu.curselection()
        if durum["rol"] in LISTE_ROLLER and sec:
            return gercek[sec[0]]
        return metin_var.get()

    def yaz(dx, dy):
        excel = durum["excel"]
        if excel is None:
            durum_var.set("Excel'e bağlı değil.")
            return "break"
        # kutu boş olsa bile listeden ↓ ile seçim yapıldıysa onu yaz
        secim = kutu.curselection() if durum["rol"] in LISTE_ROLLER else ()
        metin = secili_metin() if (metin_var.get().strip() or secim) else ""
        try:
            hucreye_yaz(excel, metin, dx, dy)
            if metin:
                durum_var.set(f"{durum['adres']} ← {metin[:40]}")
            metin_var.set("")
            izle()
        except Exception:
            durum_var.set("Yazılamadı — Excel meşgul (hücre düzenleme modunda olabilir).")
        return "break"

    def kaydir(yon):
        if durum["rol"] in LISTE_ROLLER and kutu.size():
            sec = kutu.curselection()
            if not sec:
                if yon > 0:  # ilk ↓ ile seçim başlar
                    kutu.selection_set(0)
                    kutu.see(0)
            else:
                i = sec[0] + yon
                kutu.selection_clear(0, "end")
                if i >= 0:  # en üstteyken ↑ = seçimi bırak, ham metne dön
                    i = min(kutu.size() - 1, i)
                    kutu.selection_set(i)
                    kutu.see(i)
        else:
            yaz(0, yon)  # serbest hücrede ↑↓ = yaz ve aşağı/yukarı
        return "break"

    def yeni_personel(_e=None):
        """Ctrl+N: kutudaki adı ÖNCE PERSONEL LİSTESİ'ne ekler, SONRA aktif
        hücreye yazıp sağa geçer. İlk basış onay ister; ad zaten listedeyse
        eklemeden doğru yazımıyla hücreye yazar."""
        excel = durum["excel"]
        if excel is None:
            durum_var.set("Excel'e bağlı değil.")
            return "break"
        if durum["rol"] != "PERSONEL":
            durum_var.set("Yeni personel ekleme PERSONEL sütunundayken kullanılır.")
            return "break"
        ad = tr_upper(metin_var.get())
        if len(ad.split()) < 2:
            durum_var.set("Önce AD SOYAD yaz (en az iki kelime), sonra Ctrl+N.")
            return "break"
        mevcut = next((a for a in listeler.get("PERSONEL", [])
                       if fold(a) == fold(ad)), None)
        if mevcut:
            try:
                hucreye_yaz(excel, mevcut, 1, 0)
                metin_var.set("")
                izle()
                durum_var.set(f"{mevcut} zaten listedeydi — hücreye yazıldı.")
            except Exception:
                durum_var.set("Yazılamadı — Excel meşgul.")
            return "break"
        if durum.get("yeni_onay") != fold(ad):
            durum["yeni_onay"] = fold(ad)
            durum_var.set(f"YENİ PERSONEL 『{ad}』 listeye eklenecek — onay için tekrar Ctrl+N.")
            return "break"
        durum["yeni_onay"] = None
        try:
            satir, hata = personel_listeye_ekle(excel, ad)
        except Exception as e:
            satir, hata = None, str(e)
        if hata:
            durum_var.set("Listeye eklenemedi: " + hata)
            return "break"
        listeler["PERSONEL"].append(ad)
        listeler["PERSONEL"].sort(key=fold)
        try:
            hucreye_yaz(excel, ad, 1, 0)
            metin_var.set("")
            izle()
            durum_var.set(f"➕ {ad} → PERSONEL LİSTESİ satır {satir} + hücreye yazıldı.")
        except Exception:
            durum_var.set(f"Listeye eklendi (satır {satir}) ama hücreye yazılamadı — tekrar dene.")
        return "break"

    def ctrl_d(_e=None):
        excel = durum["excel"]
        if excel is None:
            durum_var.set("Excel'e bağlı değil.")
            return "break"
        try:
            if ustu_kopyala(excel):
                durum_var.set(f"{durum['adres']} ← üstteki satırdan kopyalandı (Ctrl+D)")
            else:
                durum_var.set("En üst satırdasın, üstünde kopyalanacak hücre yok.")
            izle()
        except Exception:
            durum_var.set("Kopyalanamadı — Excel meşgul (hücre düzenleme modunda olabilir).")
        return "break"

    giris.bind("<Right>", lambda e: yaz(1, 0))
    giris.bind("<Left>", lambda e: yaz(-1, 0))
    giris.bind("<Return>", lambda e: yaz(1, 0))
    giris.bind("<Down>", lambda e: kaydir(1))
    giris.bind("<Up>", lambda e: kaydir(-1))
    giris.bind("<Escape>", lambda e: (metin_var.set(""), oner(), "break")[-1])
    giris.bind("<Control-d>", ctrl_d)
    giris.bind("<Control-D>", ctrl_d)
    root.bind("<Control-d>", ctrl_d)
    root.bind("<Control-D>", ctrl_d)
    yeni_btn.config(command=yeni_personel)
    giris.bind("<Control-n>", yeni_personel)
    giris.bind("<Control-N>", yeni_personel)
    root.bind("<Control-n>", yeni_personel)
    root.bind("<Control-N>", yeni_personel)
    giris.bind("<KeyRelease>", lambda e: None if e.keysym in
               ("Right", "Left", "Up", "Down", "Return", "Escape") else oner())
    kutu.bind("<Double-Button-1>", lambda e: yaz(1, 0))

    def izle():
        excel = durum["excel"]
        if excel is None:
            excel = excel_bul()
            if excel is None:
                baslik_var.set("Excel açık değil")
                durum_var.set("Excel'i ve stok dosyanı aç, kendiliğinden bağlanır.")
                return
            durum["excel"] = excel
        try:
            bilgi = hucre_bilgi(excel)
            if bilgi is None:
                baslik_var.set("Excel'de dosya açık değil")
                return
            sayfa, adres, satir, kolon, rol = bilgi
            durum["adres"] = adres
            degisti = (rol != durum["rol"]) or (sayfa != durum["sayfa"])
            durum["rol"], durum["sayfa"] = rol, sayfa
            baslik_var.set(f"{adres}  —  {rol}" if rol != "SERBEST"
                           else f"{adres}  —  serbest yazım")
            if degisti:
                oner()
        except Exception:
            durum["excel"] = excel_bul()
            baslik_var.set("Excel meşgul…")

    sayac = {"n": 0}

    def dongu():
        izle()
        sayac["n"] += 1
        if sayac["n"] % 20 == 0 and durum["excel"] is not None:
            # ~10 sn'de bir liman stoğunu açık Excel'den CANLI tazele
            stok = depo_stok_com(durum["excel"])
            if stok:
                listeler["MALZEME_STOK"] = stok
                if durum["rol"] == "MALZEME" and not kutu.curselection():
                    oner()
        root.after(500, dongu)

    if excel_bul() is None:
        try:
            os.startfile(DOSYA_TCH)
            durum_var.set("Excel açılıyor, dosya yüklenince bağlanacak…")
        except OSError:
            durum_var.set("TCH dosyası bulunamadı — Excel'i elle aç.")
    dongu()
    root.mainloop()


if __name__ == "__main__":
    gui()
