# -*- coding: utf-8 -*-
"""
DOLAP KARŞILAŞTIRMA — dolap almamış personel listesi çıkarma aracı.

Girdiler (bu script'in bulunduğu klasörde):
  - ZIMMET LİSTESİ.xlsm      -> DOLAP sayfası (B sütunu isimler; adı geçen dolap aldı)
  - DOĞRU PERSONEL İSİMLERİ.xlsx -> güncel personel (AD-SOYAD, TC)
  - AYRILMIŞ PERSONEL LİSTESİ.xlsx -> işten ayrılanlar (AD-SOYAD, TC)
  - SAHA PERSONELLERİ LİSTE.xlsx -> bilgi amaçlı (raporda "SAHA" sütunu)

İşleyiş:
  1) Güncel listeyle tam eşleşen -> dolap ALMIŞ (otomatik).
     Ayrılmış listeyle tam eşleşen -> AYRILMIŞ (otomatik, soru sorulmaz).
     İki listede birden olan güncel sayılır (tekrar işe girmiş olabilir).
  2) Kalan isimler SIRAYLA gösterilir: öneri listesinde güncel VE ayrılmış
     personel birlikte çıkar (ayrılmışlar [AYRILMIŞ] etiketiyle).
     ↓ ile seç + Enter = eşleştir; kutuya yazarsan canlı arama yapar.
     "Hiçbir listede yok" = iki listede de bulunamayan isim.
  3) Kararlar anında dolap_kararlar.json'a yazılır — yarıda bırakıp devam edilebilir.
  4) Bitince rapor üretilir (ALMAMIŞ + DÜZELTİLEN + AYRILMIŞ + LİSTEDE YOK + ÖZET)
     ve düzeltmeler AÇIK Excel'deki ZIMMET LİSTESİ'ne COM ile OTOMATİK uygulanır
     (makrolar bozulmaz; dosyayı kaydetme sana kalır).
"""
import os
import sys
import json
import shutil
import difflib
import tempfile
import unicodedata
import warnings
import datetime

warnings.filterwarnings("ignore")

BURA = os.path.dirname(os.path.abspath(__file__))
ZIMMET = os.path.join(BURA, "ZIMMET LİSTESİ.xlsm")
DOGRU = os.path.join(BURA, "DOĞRU PERSONEL İSİMLERİ.xlsx")
AYRILMIS = os.path.join(BURA, "AYRILMIŞ PERSONEL LİSTESİ.xlsx")
SAHA = os.path.join(BURA, "SAHA PERSONELLERİ LİSTE.xlsx")
KARAR_JSON = os.path.join(BURA, "dolap_kararlar.json")
RAPOR_XLSX = os.path.join(BURA, "DOLAP_KARSILASTIRMA_SONUC.xlsx")

_TR = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")


def fold(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s)).replace("̇", "")  # 'İ'.lower() artığı
    return " ".join(s.translate(_TR).upper().split())


def _kopyadan_ac(yol):
    from openpyxl import load_workbook
    kopya = os.path.join(tempfile.gettempdir(), f"dolap_{os.getpid()}_{os.path.basename(yol)}")
    shutil.copy2(yol, kopya)
    wb = load_workbook(kopya, read_only=True, data_only=True)
    return wb, kopya


def veri_yukle():
    dogru, tc = {}, {}
    wb, k = _kopyadan_ac(DOGRU)
    for r in wb.active.iter_rows(min_row=2, values_only=True):
        if r[0] and str(r[0]).strip():
            f = fold(r[0])
            dogru.setdefault(f, str(r[0]).strip())
            if len(r) > 1 and r[1]:
                tc.setdefault(f, r[1])
    wb.close(); os.remove(k)

    ayrilmis, ayr_tc = {}, {}
    try:
        wb, k = _kopyadan_ac(AYRILMIS)
        for r in wb.active.iter_rows(min_row=2, values_only=True):
            if r[0] and str(r[0]).strip():
                f = fold(r[0])
                if f in dogru:      # iki listede birden -> güncel sayılır
                    continue
                ayrilmis.setdefault(f, str(r[0]).strip())
                if len(r) > 1 and r[1]:
                    ayr_tc.setdefault(f, r[1])
        wb.close(); os.remove(k)
    except Exception:
        pass

    saha = set()
    try:
        wb, k = _kopyadan_ac(SAHA)
        for r in wb.active.iter_rows(min_row=2, values_only=True):
            if r[0]:
                saha.add(fold(r[0]))
        wb.close(); os.remove(k)
    except Exception:
        pass

    dolap = []  # (satır, dolap no, isim)
    wb, k = _kopyadan_ac(ZIMMET)
    for i, r in enumerate(wb["DOLAP"].iter_rows(min_row=2, values_only=True), start=2):
        if r[1] and str(r[1]).strip():
            dolap.append({"satir": i, "no": r[0], "ad": str(r[1]).strip()})
    wb.close(); os.remove(k)
    return dogru, tc, saha, ayrilmis, ayr_tc, dolap


def kararlari_oku():
    try:
        with open(KARAR_JSON, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def kararlari_yaz(kararlar):
    gecici = KARAR_JSON + ".tmp"
    with open(gecici, "w", encoding="utf-8") as f:
        json.dump(kararlar, f, ensure_ascii=False, indent=1)
    os.replace(gecici, KARAR_JSON)


def anahtar(kayit):
    return f"{kayit['satir']}|{fold(kayit['ad'])}"


# ------------------------------------------------------------------ ortak sınıflandırma
def siniflandir(dogru, ayrilmis, dolap, kararlar):
    """Her dolap kaydını sınıflar. Döner: alanlar (fold set), duzeltmeler,
    ayrilanlar [(satır, no, dolaptaki ad, listedeki ad, tür)], yok, kararsiz."""
    alanlar = set()
    duzeltmeler = []         # (satır, dolap no, eski, yeni, tür)
    ayrilanlar = []
    yok = []
    kararsiz = []
    for d in dolap:
        f = fold(d["ad"])
        kr = kararlar.get(anahtar(d))
        if f in dogru:
            alanlar.add(f)
            if d["ad"] != dogru[f]:
                duzeltmeler.append((d["satir"], d["no"], d["ad"], dogru[f], "OTOMATİK (yazım)"))
        elif kr and kr["karar"] == "ESLE" and kr["hedef"] in dogru:
            alanlar.add(kr["hedef"])
            duzeltmeler.append((d["satir"], d["no"], d["ad"], dogru[kr["hedef"]], "ELLE EŞLEŞTİRME"))
        elif kr and kr["karar"] == "AYRILDI" and kr["hedef"] in ayrilmis:
            ayrilanlar.append((d["satir"], d["no"], d["ad"], ayrilmis[kr["hedef"]], "ELLE (ayrılmış listesi)"))
            if d["ad"] != ayrilmis[kr["hedef"]]:
                duzeltmeler.append((d["satir"], d["no"], d["ad"], ayrilmis[kr["hedef"]], "ELLE (AYRILMIŞ yazım)"))
        elif f in ayrilmis:
            # ayrılmış listesiyle birebir eşleşme, eski "YOK" kararından üstündür
            ayrilanlar.append((d["satir"], d["no"], d["ad"], ayrilmis[f], "OTOMATİK (ayrılmış listesi)"))
            if d["ad"] != ayrilmis[f]:
                duzeltmeler.append((d["satir"], d["no"], d["ad"], ayrilmis[f], "OTOMATİK (ayrılmış yazım)"))
        elif kr and kr["karar"] == "YOK":
            yok.append((d["satir"], d["no"], d["ad"]))
        else:
            kararsiz.append(d)
    return alanlar, duzeltmeler, ayrilanlar, yok, kararsiz


# ------------------------------------------------------------------ rapor
def rapor_yaz(dogru, tc, saha, ayrilmis, ayr_tc, dolap, kararlar):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    alanlar, duzeltmeler, ayrilanlar, yok, kararsiz = siniflandir(dogru, ayrilmis, dolap, kararlar)
    almamis = [f for f in dogru if f not in alanlar]

    wb = Workbook()

    def sayfa(ad, basliklar, satirlar, genislikler):
        ws = wb.create_sheet(ad) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = ad
        ws.append(basliklar)
        for c in ws[1]:
            c.font = Font(name="Arial", bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="1F4E78")
        for r in satirlar:
            ws.append(r)
        for i, g in enumerate(genislikler, 1):
            ws.column_dimensions[get_column_letter(i)].width = g
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(basliklar))}{ws.max_row}"
        return ws

    sayfa("DOLAP ALMAMIŞ",
          ["AD-SOYAD", "TC", "SAHA LİSTESİNDE"],
          sorted(([dogru[f], tc.get(f, ""), "EVET" if f in saha else ""] for f in almamis),
                 key=lambda r: fold(r[0])),
          [32, 16, 16])
    sayfa("DÜZELTİLENLER",
          ["DOLAP SATIRI", "DOLAP NO", "ESKİ İSİM", "DOĞRU İSİM", "TÜR"],
          duzeltmeler, [13, 10, 30, 30, 26])
    sayfa("AYRILMIŞ",
          ["DOLAP SATIRI", "DOLAP NO", "DOLAPTAKİ İSİM", "LİSTEDEKİ İSİM", "TC", "TÜR"],
          [(s, n, a, la, ayr_tc.get(fold(la), ""), t) for s, n, a, la, t in ayrilanlar],
          [13, 10, 30, 30, 15, 26])
    sayfa("LİSTEDE YOK",
          ["DOLAP SATIRI", "DOLAP NO", "İSİM"],
          yok, [13, 10, 32])
    if kararsiz:
        sayfa("KARARSIZ",
              ["DOLAP SATIRI", "DOLAP NO", "İSİM"],
              [(d["satir"], d["no"], d["ad"]) for d in kararsiz], [13, 10, 32])
    sayfa("ÖZET", ["BİLGİ", "SAYI"], [
        ("Güncel personel (DOĞRU liste)", len(dogru)),
        ("Ayrılmış personel listesi", len(ayrilmis)),
        ("DOLAP sayfasında isimli satır", len(dolap)),
        ("Dolap almış (eşleşen kişi)", len(alanlar)),
        ("DOLAP ALMAMIŞ", len(almamis)),
        ("Düzeltilen isim", len(duzeltmeler)),
        ("Ayrılmış (dolap kaydı)", len(ayrilanlar)),
        ("Hiçbir listede yok", len(yok)),
        ("Kararsız (tamamlanmadı)", len(kararsiz)),
        ("Rapor tarihi", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")),
    ], [34, 22])

    gecici = RAPOR_XLSX + ".tmp"
    wb.save(gecici)
    os.replace(gecici, RAPOR_XLSX)
    return len(almamis), len(duzeltmeler), len(ayrilanlar) + len(yok), len(kararsiz)


def duzeltmeleri_uygula(dogru, ayrilmis, dolap, kararlar):
    """Düzeltilmiş isimleri (güncel + ayrılmış eşleşmeleri dahil) AÇIK Excel'deki
    ZIMMET LİSTESİ.xlsm DOLAP sayfasına COM ile yazar (makrolar korunur).
    Kaydetmez — kontrol edip sen kaydedersin."""
    import win32com.client
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return None, "Excel açık değil. Önce ZIMMET LİSTESİ.xlsm'i Excel'de aç."
    hedef = None
    for w in excel.Workbooks:
        if fold(w.Name) == fold(os.path.basename(ZIMMET)):
            hedef = w
            break
    if hedef is None:
        return None, "ZIMMET LİSTESİ.xlsm açık Excel'de bulunamadı — dosyayı aç ve tekrar dene."
    ws = hedef.Worksheets("DOLAP")
    _, duzeltmeler, _, _, _ = siniflandir(dogru, ayrilmis, dolap, kararlar)
    yazilan = 0
    for satir, no, eski, yeni, tur in duzeltmeler:
        mevcut = ws.Cells(satir, 2).Value
        if mevcut is not None and fold(mevcut) == fold(eski):
            ws.Cells(satir, 2).Value = yeni
            yazilan += 1
    return yazilan, None


# ------------------------------------------------------------------ GUI
def gui():
    import tkinter as tk
    from tkinter import ttk, messagebox

    for yol in (ZIMMET, DOGRU):
        if not os.path.exists(yol):
            r = tk.Tk(); r.withdraw()
            messagebox.showerror("Dosya yok", f"Bulunamadı:\n{yol}\n\n"
                                 "Bu script, üç Excel'in olduğu klasörde durmalı.")
            return

    dogru, tc, saha, ayrilmis, ayr_tc, dolap = veri_yukle()
    kararlar = kararlari_oku()
    # öneri havuzu: güncel + ayrılmış birlikte (fold -> (görünen ad, kaynak))
    havuz = {k: (v, "AYRILMIŞ") for k, v in ayrilmis.items()}
    havuz.update({k: (v, "GÜNCEL") for k, v in dogru.items()})
    anahtar_liste = list(havuz)

    tam_sayi = sum(1 for d in dolap if fold(d["ad"]) in dogru)
    oto_ayr = sum(1 for d in dolap
                  if fold(d["ad"]) not in dogru and fold(d["ad"]) in ayrilmis
                  and anahtar(d) not in kararlar)
    kuyruk = [d for d in dolap
              if fold(d["ad"]) not in dogru and fold(d["ad"]) not in ayrilmis
              and anahtar(d) not in kararlar]

    root = tk.Tk()
    root.title("DOLAP KARŞILAŞTIRMA — İSİM EŞLEŞTİRME")
    root.geometry("780x640")
    stil = ttk.Style(root)
    try:
        stil.theme_use("vista")
    except Exception:
        pass

    durum = {"i": 0, "gecmis": []}

    ust_var = tk.StringVar()
    ttk.Label(root, textvariable=ust_var, font=("Segoe UI", 11),
              padding=(14, 8, 14, 0)).pack(fill="x")
    ad_var = tk.StringVar()
    ttk.Label(root, textvariable=ad_var, font=("Segoe UI", 20, "bold"),
              foreground="#B33", padding=(14, 2)).pack(fill="x")
    bilgi_var = tk.StringVar()
    ttk.Label(root, textvariable=bilgi_var, font=("Segoe UI", 10),
              foreground="#555", padding=(14, 0)).pack(fill="x")

    ara_cerceve = ttk.Frame(root, padding=(14, 8, 14, 2))
    ara_cerceve.pack(fill="x")
    ttk.Label(ara_cerceve, text="Ara:", font=("Segoe UI", 12)).pack(side="left")
    ara_var = tk.StringVar()
    ara = ttk.Entry(ara_cerceve, textvariable=ara_var, font=("Segoe UI", 14))
    ara.pack(side="left", fill="x", expand=True, padx=(8, 0))

    kutu = tk.Listbox(root, font=("Segoe UI", 13), height=10,
                      selectbackground="#1F4E78", selectforeground="white")
    kutu.pack(fill="both", expand=True, padx=14, pady=(4, 2))
    gercek = []

    ipucu = ("↓ ↑ listeden seç   Enter = EŞLEŞTİR   |   kutuya yazınca canlı arama\n"
             "[AYRILMIŞ] etiketli seçim = işten ayrılmış eşleşmesi   |   "
             "sağdaki düğme: hiçbir listede yok")
    ttk.Label(root, text=ipucu, font=("Segoe UI", 10), foreground="#555",
              padding=(14, 2)).pack(fill="x")

    dugme = ttk.Frame(root, padding=(14, 4, 14, 10))
    dugme.pack(fill="x")
    esle_btn = tk.Button(dugme, text="✔ EŞLEŞTİR (Enter)", bg="#2e7d32", fg="white",
                         font=("Segoe UI", 12, "bold"), padx=14)
    esle_btn.pack(side="left")
    yok_btn = tk.Button(dugme, text="✖ Hiçbir listede yok", bg="#b33", fg="white",
                        font=("Segoe UI", 12, "bold"), padx=14)
    yok_btn.pack(side="left", padx=10)
    geri_btn = tk.Button(dugme, text="↶ Geri", font=("Segoe UI", 11), padx=10)
    geri_btn.pack(side="left")

    alt_var = tk.StringVar()
    ttk.Label(root, textvariable=alt_var, relief="sunken", anchor="w",
              padding=(10, 4), font=("Segoe UI", 10)).pack(fill="x", side="bottom")

    def oneri_doldur(kayit):
        kutu.delete(0, "end")
        gercek.clear()
        aranan = fold(ara_var.get())
        if aranan:
            kelimeler = aranan.split()
            adaylar = [k for k in anahtar_liste if all(x in k for x in kelimeler)][:40]
        else:
            f = fold(kayit["ad"])
            adaylar = difflib.get_close_matches(f, anahtar_liste, n=8, cutoff=0.5)
            kel = set(f.split())
            for k in anahtar_liste:      # ortak kelimesi olanları da ekle
                if k not in adaylar and kel & set(k.split()):
                    adaylar.append(k)
            adaylar = adaylar[:15]
        f0 = fold(kayit["ad"])
        oranlar = []
        for k in adaylar:
            gercek.append(k)
            ad, kaynak = havuz[k]
            oran = int(difflib.SequenceMatcher(None, f0, k).ratio() * 100)
            oranlar.append(oran)
            etiket = "  [AYRILMIŞ]" if kaynak == "AYRILMIŞ" else ""
            kutu.insert("end", f"{ad}{etiket}    (%{oran} benzer)")
        # güvenlik: ilk öneri ancak ÇOK benzerse hazır seçili gelir —
        # bariz yazım hatasında Enter yeter, şüpheli durumda bilerek seçmek gerekir
        if kutu.size() and not aranan and oranlar[0] >= 80:
            kutu.selection_set(0)

    def goster():
        if durum["i"] >= len(kuyruk):
            bitir()
            return
        kayit = kuyruk[durum["i"]]
        kalan = len(kuyruk) - durum["i"]
        ust_var.set(f"Karar {durum['i'] + 1} / {len(kuyruk)}   "
                    f"(otomatik: {tam_sayi} güncel + {oto_ayr} ayrılmış, kalan: {kalan})")
        ad_var.set(kayit["ad"])
        bilgi_var.set(f"DOLAP sayfası satır {kayit['satir']}  •  Dolap No: {kayit['no']}")
        ara_var.set("")
        oneri_doldur(kayit)
        ara.focus_set()

    def karar_ver(tip):
        kayit = kuyruk[durum["i"]]
        if tip == "ESLE":
            sec = kutu.curselection()
            if not sec:
                alt_var.set("Önce listeden bir isim seç (↓ tuşu) ya da arayıp seç.")
                return
            hedef = gercek[sec[0]]
            ad, kaynak = havuz[hedef]
            if kaynak == "AYRILMIŞ":
                kararlar[anahtar(kayit)] = {"karar": "AYRILDI", "hedef": hedef,
                                            "hedef_ad": ad}
                alt_var.set(f"{kayit['ad']}  →  {ad}  [AYRILMIŞ]")
            else:
                kararlar[anahtar(kayit)] = {"karar": "ESLE", "hedef": hedef,
                                            "hedef_ad": ad}
                alt_var.set(f"{kayit['ad']}  →  {ad}")
        else:
            kararlar[anahtar(kayit)] = {"karar": "YOK"}
            alt_var.set(f"{kayit['ad']}  →  hiçbir listede yok")
        kararlari_yaz(kararlar)
        durum["gecmis"].append(durum["i"])
        durum["i"] += 1
        goster()

    def geri(*_):
        if not durum["gecmis"]:
            return
        durum["i"] = durum["gecmis"].pop()
        kararlar.pop(anahtar(kuyruk[durum["i"]]), None)
        kararlari_yaz(kararlar)
        goster()

    def kaydir(yon):
        if not kutu.size():
            return "break"
        sec = kutu.curselection()
        i = (sec[0] + yon) if sec else 0
        i = max(0, min(kutu.size() - 1, i))
        kutu.selection_clear(0, "end")
        kutu.selection_set(i)
        kutu.see(i)
        return "break"

    def bitir():
        for w in root.winfo_children():
            w.destroy()
        ttk.Label(root, text="Eşleştirme tamamlandı ✔", font=("Segoe UI", 18, "bold"),
                  padding=(16, 16)).pack()
        uygulama_var = tk.StringVar()
        ttk.Label(root, textvariable=uygulama_var, font=("Segoe UI", 11),
                  foreground="#2e7d32", padding=(16, 2)).pack()
        ozet_var = tk.StringVar()
        ttk.Label(root, textvariable=ozet_var, font=("Segoe UI", 12),
                  padding=(16, 4)).pack()

        def raporla():
            a, dz, ay, ks = rapor_yaz(dogru, tc, saha, ayrilmis, ayr_tc, dolap, kararlar)
            ozet_var.set(f"DOLAP ALMAMIŞ: {a} kişi   |   düzeltilen isim: {dz}\n"
                         f"ayrılmış + listede yok: {ay}   |   kararsız: {ks}\n\n"
                         f"Rapor: {os.path.basename(RAPOR_XLSX)}")
            os.startfile(RAPOR_XLSX)

        def uygula(sessiz=False):
            n, hata = duzeltmeleri_uygula(dogru, ayrilmis, dolap, kararlar)
            if hata:
                uygulama_var.set("⚠ Düzeltmeler uygulanamadı: " + hata)
                if not sessiz:
                    messagebox.showerror("Uygulanamadı", hata)
            else:
                uygulama_var.set(f"✏ {n} isim ZIMMET LİSTESİ → DOLAP sayfasında düzeltildi "
                                 "(dosya kaydedilmedi — kontrol edip Ctrl+S ile kaydet).")

        tk.Button(root, text="📄 Rapor oluştur ve aç", command=raporla, bg="#1F4E78",
                  fg="white", font=("Segoe UI", 13, "bold"), padx=16, pady=6).pack(pady=8)
        tk.Button(root, text="✏ Düzeltmeleri yeniden uygula (Excel açıkken)",
                  command=uygula, font=("Segoe UI", 11), padx=12).pack(pady=4)
        ttk.Label(root, text="Kararları değiştirmek istersen dolap_kararlar.json'u silip"
                             " aracı yeniden aç.", font=("Segoe UI", 9),
                  foreground="#777", padding=(10, 10)).pack()
        # seçilenler otomatik düzeltilsin: bitişte kendiliğinden uygula
        root.after(300, lambda: uygula(sessiz=True))

    esle_btn.config(command=lambda: karar_ver("ESLE"))
    yok_btn.config(command=lambda: karar_ver("YOK"))
    geri_btn.config(command=geri)
    ara.bind("<Return>", lambda e: karar_ver("ESLE"))
    ara.bind("<Down>", lambda e: kaydir(1))
    ara.bind("<Up>", lambda e: kaydir(-1))
    ara.bind("<KeyRelease>", lambda e: None if e.keysym in
             ("Return", "Up", "Down", "Escape") else
             oneri_doldur(kuyruk[durum["i"]]) if durum["i"] < len(kuyruk) else None)
    ara.bind("<Escape>", lambda e: (ara_var.set(""),
             oneri_doldur(kuyruk[durum["i"]]) if durum["i"] < len(kuyruk) else None))
    kutu.bind("<Double-Button-1>", lambda e: karar_ver("ESLE"))

    goster()
    root.mainloop()


if __name__ == "__main__":
    gui()
