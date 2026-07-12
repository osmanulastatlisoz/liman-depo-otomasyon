# -*- coding: utf-8 -*-
"""
DOLAP VERİLECEK PERSONELLER listesi üretir.

SAHA PERSONELLERİ LİSTE.xlsx'ten, ZIMMET LİSTESİ DOLAP sayfasında adı geçen
(dolap almış) kişileri çıkarır; kalanları işe giriş tarihleriyle birlikte
"DOLAP VERİLECEK PERSONELLER.xlsx" dosyasına yazar.

Dolap almış tespiti dolap_eslestir ile aynı mantıktır: birebir eşleşmeler +
senin elle eşleştirme kararların (dolap_kararlar.json). Kararsız isim kaldıysa
uyarır — kararları bitirip yeniden çalıştırınca liste netleşir.

Kullanım: python dolap_verilecek_liste.py  (çıktıyı da açar)
"""
import os
import sys
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dolap_eslestir as de

CIKTI = os.path.join(de.BURA, "DOLAP VERİLECEK PERSONELLER.xlsx")


def uret(ac=False):
    dogru, tc, saha_set, ayrilmis, ayr_tc, dolap = de.veri_yukle()
    kararlar = de.kararlari_oku()
    alanlar, _dz, _ayr, _yok, kararsiz = de.siniflandir(dogru, ayrilmis, dolap, kararlar)

    # saha listesini tarihiyle oku
    saha = []
    wb, k = de._kopyadan_ac(de.SAHA)
    for r in wb.active.iter_rows(min_row=2, values_only=True):
        if r[0] and str(r[0]).strip():
            saha.append({"ad": str(r[0]).strip(),
                         "tc": r[1] if len(r) > 1 else None,
                         "giris": r[2] if len(r) > 2 else None})
    wb.close(); os.remove(k)

    verilecek = [s for s in saha if de.fold(s["ad"]) not in alanlar]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "DOLAP VERİLECEK"
    ws.append(["SIRA", "AD-SOYAD", "TC", "İŞE GİRİŞ TARİHİ"])
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.alignment = Alignment(horizontal="center")
    for i, s in enumerate(sorted(verilecek, key=lambda x: de.fold(x["ad"])), 1):
        ws.append([i, s["ad"], s["tc"], s["giris"]])
        ws.cell(row=ws.max_row, column=4).number_format = "DD.MM.YYYY"
    for col, w in zip("ABCD", [7, 34, 16, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    gecici = CIKTI + ".tmp"
    wb.save(gecici)
    os.replace(gecici, CIKTI)

    if ac:
        os.startfile(CIKTI)
    return len(saha), len(saha) - len(verilecek), len(verilecek), len(kararsiz)


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    toplam, almis, verilecek, kararsiz = uret(ac="--acma" not in sys.argv)
    print(f"Saha personeli: {toplam}")
    print(f"Dolap almış (çıkarılan): {almis}")
    print(f"DOLAP VERİLECEK: {verilecek} kişi -> {os.path.basename(CIKTI)}")
    if kararsiz:
        print(f"UYARI: {kararsiz} isim hâlâ kararsız (dolap_eslestir'de bitirilmedi) — "
              "kararları bitirip bunu yeniden çalıştır.")
