# -*- coding: utf-8 -*-
"""
BİRLEŞİK STOK SORGULAMA
- 2026 DEPO STOK.xlsm  -> STOK GİRİŞ ÇIKIŞ
- TCH_STOK_TAKIP_CALISMASI GÜNCEL.xlsm -> STOK HAREKETLERİ
İkisini birleştirir (birebir mükerrerleri teke indirir), formülsüz
"BİRLEŞİK STOK HAREKETLERİ.xlsx" dosyasını günceller ve personel adına
göre anında sorgulama penceresi açar.

Kullanım:  çift tık (PERSONEL SORGU.bat)  veya  python birlesik_stok_sorgu.py --rebuild
"""
import os
import sys
import json
import difflib
import pickle
import queue
import shutil
import tempfile
import threading
import unicodedata
import warnings
import datetime
from collections import Counter

warnings.filterwarnings("ignore")

MASAUSTU = os.path.join(os.path.expanduser("~"), "Desktop")
AYAR_JSON = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
                         "DepoAraclari", "yollar.json")


def kayitli_yol(anahtar, varsayilanlar, baslik):
    """Dosyayı sırayla arar: kayıtlı ayar -> varsayılan konumlar -> kullanıcıya
    sor. Seçim yollar.json'a kaydedilir, bir daha sorulmaz — böylece araç
    farklı klasör düzenine sahip bilgisayarlarda da çalışır."""
    try:
        with open(AYAR_JSON, encoding="utf-8") as f:
            kayit = json.load(f)
    except Exception:
        kayit = {}
    yol = kayit.get(anahtar)
    if yol and os.path.exists(yol):
        return yol
    for v in varsayilanlar:
        if os.path.exists(v):
            return v
    import tkinter as tk
    from tkinter import filedialog
    kok = tk.Tk(); kok.withdraw()
    yol = filedialog.askopenfilename(
        title=baslik, filetypes=[("Excel", "*.xlsm *.xlsx"), ("Tümü", "*.*")])
    kok.destroy()
    if not yol:
        raise SystemExit(f"{baslik} — dosya seçilmedi, çıkılıyor.")
    kayit[anahtar] = yol
    try:
        os.makedirs(os.path.dirname(AYAR_JSON), exist_ok=True)
        with open(AYAR_JSON, "w", encoding="utf-8") as f:
            json.dump(kayit, f, ensure_ascii=False, indent=1)
    except OSError:
        pass
    return yol


DOSYA_2026 = kayitli_yol("depo_stok_2026",
                         [os.path.join(MASAUSTU, "2026 DEPO STOK.xlsm")],
                         "2026 DEPO STOK.xlsm dosyasını seç")
DOSYA_TCH = kayitli_yol("tch_stok_takip",
                        [os.path.join(MASAUSTU, "TCH_STOK_TAKIP_CALISMASI GÜNCEL.xlsm")],
                        "TCH_STOK_TAKIP_CALISMASI GÜNCEL.xlsm dosyasını seç")
CIKTI_XLSX = os.path.join(os.path.dirname(DOSYA_TCH) or MASAUSTU,
                          "BİRLEŞİK STOK HAREKETLERİ.xlsx")
CACHE_DIR = os.path.join(os.environ.get("LOCALAPPDATA", MASAUSTU), "BirlesikStok")
CACHE_PKL = os.path.join(CACHE_DIR, "birlesik_cache.pkl")

BASLIKLAR = ["TARİH", "İŞLEM TÜRÜ", "PERSONEL", "MALZEME ADI", "MİKTAR",
             "AÇIKLAMA / PLAKA", "MALZEME KODU", "MALZEME TÜRÜ", "HEDEF DEPO", "KAYNAK"]

_TR = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")


def fold(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return " ".join(s.translate(_TR).upper().split())


def _tarih(v):
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day)
    return None


def _miktar_key(m):
    try:
        return float(m)
    except (TypeError, ValueError):
        return fold(m)


def _anahtar(r):
    t = r[0].date() if r[0] else None
    return (t, fold(r[1]), fold(r[2]), fold(r[3]), _miktar_key(r[4]))


def _kopyadan_ac(yol):
    """Kaynak Excel'de açıkken bile güvenle okumak için geçici kopyadan açar;
    asıl dosyaya yalnızca kopyalama anında (< 1 sn) dokunulur."""
    from openpyxl import load_workbook
    kopya = os.path.join(tempfile.gettempdir(),
                         f"bss_{os.getpid()}_{os.path.basename(yol)}")
    shutil.copy2(yol, kopya)
    return load_workbook(kopya, read_only=True, data_only=True), kopya


def oku_2026():
    wb, kopya = _kopyadan_ac(DOSYA_2026)
    try:
        ws = wb["STOK GİRİŞ ÇIKIŞ"]
        satirlar = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            islem, personel, malzeme, miktar = r[0], r[1], r[2], r[3]
            if not malzeme and not personel:
                continue
            satirlar.append((_tarih(r[6]), islem, personel, malzeme, miktar,
                             r[4], r[8], r[9], None, "2026 DEPO STOK"))
        return satirlar
    finally:
        wb.close()
        try:
            os.remove(kopya)
        except OSError:
            pass


def oku_tch():
    wb, kopya = _kopyadan_ac(DOSYA_TCH)
    try:
        ws = wb["STOK HAREKETLERİ"]
        satirlar = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            tarih, depo, islem, personel, malzeme, miktar = r[0], r[1], r[2], r[3], r[4], r[5]
            if not malzeme and not personel:
                continue
            satirlar.append((_tarih(tarih), islem, personel, malzeme, miktar,
                             r[6], r[7], r[11], depo, "TCH STOK TAKİP"))
        return satirlar
    finally:
        wb.close()
        try:
            os.remove(kopya)
        except OSError:
            pass


def birlestir():
    """TCH esas alınır; 2026'daki satırlardan TCH'de birebir karşılığı
    (tarih+işlem+personel+malzeme+miktar) olanlar teke indirilir."""
    tch = oku_tch()
    eski = oku_2026()
    tch_sayac = Counter(_anahtar(r) for r in tch)
    gorulen = Counter()
    birlesik = list(tch)
    for r in eski:
        k = _anahtar(r)
        gorulen[k] += 1
        if gorulen[k] > tch_sayac[k]:
            birlesik.append(r)
    birlesik.sort(key=lambda r: (r[0] or datetime.datetime.min, r[9]))
    return birlesik


def xlsx_yaz(satirlar, hedef=CIKTI_XLSX):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "BİRLEŞİK HAREKETLER"
    ws.append(BASLIKLAR)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r in satirlar:
        ws.append(r)
        hucre = ws.cell(row=ws.max_row, column=1)
        hucre.number_format = "DD.MM.YYYY"
    genislikler = [12, 12, 26, 42, 9, 24, 15, 22, 14, 16]
    for i, g in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(i)].width = g
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    gecici = hedef + ".tmp"
    wb.save(gecici)
    os.replace(gecici, hedef)


def cache_yaz(satirlar):
    os.makedirs(CACHE_DIR, exist_ok=True)
    gecici = CACHE_PKL + ".tmp"
    with open(gecici, "wb") as f:
        pickle.dump({"zaman": datetime.datetime.now(), "satirlar": satirlar}, f)
    os.replace(gecici, CACHE_PKL)


def cache_oku():
    try:
        with open(CACHE_PKL, "rb") as f:
            d = pickle.load(f)
        return d["satirlar"], d["zaman"]
    except Exception:
        return None, None


def tam_guncelle():
    satirlar = birlestir()
    excel_hata = None
    try:
        xlsx_yaz(satirlar)
    except PermissionError:
        excel_hata = "Excel dosyası açık olduğu için dosya güncellenemedi (ekrandaki veriler günceldir)."
    cache_yaz(satirlar)
    return satirlar, excel_hata


# ------------------------------------------------------------------ GUI
def gui():
    import tkinter as tk
    from tkinter import ttk, messagebox

    for d in (DOSYA_2026, DOSYA_TCH):
        if not os.path.exists(d):
            root = tk.Tk(); root.withdraw()
            messagebox.showerror("Dosya bulunamadı", f"Kaynak dosya bulunamadı:\n{d}")
            return

    root = tk.Tk()
    root.title("PERSONEL MALZEME SORGULAMA")
    root.geometry("1150x680")
    root.minsize(900, 500)

    stil = ttk.Style(root)
    try:
        stil.theme_use("vista")
    except Exception:
        pass
    stil.configure("Treeview", font=("Segoe UI", 11), rowheight=26)
    stil.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"))

    veri = {"satirlar": [], "zaman": None}
    kuyruk = queue.Queue()

    # ---- üst şerit: arama kutuları
    ust = ttk.Frame(root, padding=(12, 10, 12, 2))
    ust.pack(fill="x")
    ttk.Label(ust, text="Personel Adı:", font=("Segoe UI", 13, "bold")).pack(side="left")
    ad_var = tk.StringVar()
    ad_giris = ttk.Entry(ust, textvariable=ad_var, font=("Segoe UI", 15), width=22)
    ad_giris.pack(side="left", padx=(8, 20))
    ad_giris.focus_set()

    ttk.Label(ust, text="Malzeme Adı:", font=("Segoe UI", 13, "bold")).pack(side="left")
    malzeme_var = tk.StringVar()
    malzeme_giris = ttk.Entry(ust, textvariable=malzeme_var, font=("Segoe UI", 15), width=22)
    malzeme_giris.pack(side="left", padx=(8, 16))

    def temizle():
        ad_var.set("")
        malzeme_var.set("")
        ad_giris.focus_set()
        filtrele()

    ttk.Button(ust, text="Temizle", command=temizle).pack(side="left")

    # ---- ikinci şerit: filtreler
    filtre = ttk.Frame(root, padding=(12, 2, 12, 4))
    filtre.pack(fill="x")
    ttk.Label(filtre, text="İşlem:", font=("Segoe UI", 11)).pack(side="left")
    islem_var = tk.StringVar(value="ÇIKIŞ")
    islem_cb = ttk.Combobox(filtre, textvariable=islem_var, state="readonly",
                            width=10, font=("Segoe UI", 11))
    islem_cb.pack(side="left", padx=(4, 16))

    ttk.Label(filtre, text="Görünüm:", font=("Segoe UI", 11)).pack(side="left")
    gorunum_var = tk.StringVar(value="Hareket listesi")
    gorunum_cb = ttk.Combobox(filtre, textvariable=gorunum_var, state="readonly", width=17,
                              values=["Hareket listesi", "Malzeme toplamları"], font=("Segoe UI", 11))
    gorunum_cb.pack(side="left", padx=(4, 16))

    ttk.Button(filtre, text="Excel'i Aç",
               command=lambda: os.startfile(CIKTI_XLSX) if os.path.exists(CIKTI_XLSX) else None
               ).pack(side="right")

    bilgi_var = tk.StringVar(value="Personel adı ve/veya malzeme adı yazın…")
    ttk.Label(root, textvariable=bilgi_var, font=("Segoe UI", 11, "bold"),
              padding=(14, 2)).pack(fill="x")

    # ---- tablo
    orta = ttk.Frame(root, padding=(12, 2, 12, 4))
    orta.pack(fill="both", expand=True)
    agac = ttk.Treeview(orta, show="headings", selectmode="browse")
    dikey = ttk.Scrollbar(orta, orient="vertical", command=agac.yview)
    agac.configure(yscrollcommand=dikey.set)
    agac.pack(side="left", fill="both", expand=True)
    dikey.pack(side="right", fill="y")
    agac.tag_configure("tek", background="#F2F7FC")

    HAREKET_KOLON = [("TARİH", 95), ("İŞLEM", 70), ("PERSONEL", 190), ("MALZEME", 330),
                     ("MİKTAR", 70), ("AÇIKLAMA / PLAKA", 180), ("KAYNAK", 130)]
    OZET_KOLON = [("MALZEME", 380), ("İŞLEM", 80), ("TOPLAM MİKTAR", 120),
                  ("KAÇ KEZ", 80), ("İLK TARİH", 100), ("SON TARİH", 100)]

    def kolonlari_kur(kolonlar):
        agac.delete(*agac.get_children())
        agac["columns"] = [k for k, _ in kolonlar]
        for k, w in kolonlar:
            agac.heading(k, text=k)
            agac.column(k, width=w, anchor="w",
                        stretch=(k in ("MALZEME", "AÇIKLAMA / PLAKA")))

    def gt(t):
        return t.strftime("%d.%m.%Y") if t else ""

    def oneri_metni(kelimeler, aranan, alan):
        """alan: 2=personel, 3=malzeme — en yakın 5 adı önerir."""
        adaylar = sorted({str(r[alan]).strip() for r in veri["satirlar"] if r[alan]})
        oneri = [a for a in adaylar if any(k in fold(a) for k in kelimeler)][:5]
        if not oneri:
            fold_map = {}
            for a in adaylar:
                fold_map.setdefault(fold(a), a)
            yakin = difflib.get_close_matches(aranan, list(fold_map), n=5, cutoff=0.55)
            oneri = [fold_map[y] for y in yakin]
        return f"  Şunu mu demek istediniz: {', '.join(oneri)}?" if oneri else ""

    def filtrele(*_):
        ad_aranan = fold(ad_var.get())
        malz_aranan = fold(malzeme_var.get())
        islem_sec = islem_var.get()
        satirlar = veri["satirlar"]
        ad_kel = ad_aranan.split() if len(ad_aranan) >= 2 else []
        malz_kel = malz_aranan.split() if len(malz_aranan) >= 2 else []
        if not ad_kel and not malz_kel:
            kolonlari_kur(HAREKET_KOLON)
            bilgi_var.set("Personel adı ve/veya malzeme adı yazın… (en az 2 harf)")
            return
        if not satirlar:
            kolonlari_kur(HAREKET_KOLON)
            bilgi_var.set("Veriler yükleniyor, lütfen birkaç saniye bekleyin…")
            return
        # kelime bazlı arama: yazılan her kelime ilgili alanda geçsin (sıra önemsiz)
        isimle = [r for r in satirlar
                  if all(k in fold(r[2]) for k in ad_kel)
                  and all(k in fold(r[3]) for k in malz_kel)]
        bulunan = isimle
        if islem_sec != "TÜMÜ":
            bulunan = [r for r in isimle if fold(r[1]) == fold(islem_sec)]
        kisiler = sorted({str(r[2]).strip() for r in bulunan if r[2]})

        if not bulunan:
            kolonlari_kur(HAREKET_KOLON)
            if isimle:
                dagilim = Counter(str(r[1]).strip().upper() for r in isimle if r[1])
                dok = ", ".join(f"{v} {k}" for k, v in dagilim.most_common())
                bilgi_var.set(f"Kayıt var ama '{islem_sec}' türünde değil ({dok}) — "
                              f"İşlem filtresini TÜMÜ yapın.")
                return
            ad_var_mi = (not ad_kel) or any(all(k in fold(r[2]) for k in ad_kel) for r in satirlar)
            malz_var_mi = (not malz_kel) or any(all(k in fold(r[3]) for k in malz_kel) for r in satirlar)
            if ad_kel and not ad_var_mi:
                bilgi_var.set("Personel bulunamadı." + oneri_metni(ad_kel, ad_aranan, 2))
            elif malz_kel and not malz_var_mi:
                bilgi_var.set("Malzeme bulunamadı." + oneri_metni(malz_kel, malz_aranan, 3))
            else:
                bilgi_var.set("İkisi de var ama bu personelin bu malzemeyle hiç kaydı yok.")
            return

        if gorunum_var.get() == "Malzeme toplamları":
            kolonlari_kur(OZET_KOLON)
            gruplar = {}
            for r in bulunan:
                k = (fold(r[3]), fold(r[1]))
                g = gruplar.setdefault(k, {"ad": r[3], "islem": r[1], "top": 0.0,
                                           "kez": 0, "ilk": None, "son": None})
                try:
                    g["top"] += float(r[4] or 0)
                except (TypeError, ValueError):
                    pass
                g["kez"] += 1
                if r[0]:
                    g["ilk"] = min(g["ilk"] or r[0], r[0])
                    g["son"] = max(g["son"] or r[0], r[0])
            sirali = sorted(gruplar.values(),
                            key=lambda g: g["son"] or datetime.datetime.min, reverse=True)
            for i, g in enumerate(sirali):
                top = int(g["top"]) if g["top"] == int(g["top"]) else round(g["top"], 2)
                agac.insert("", "end", tags=("tek",) if i % 2 else (),
                            values=(g["ad"], g["islem"], top, g["kez"], gt(g["ilk"]), gt(g["son"])))
            bilgi_var.set(f"{len(kisiler)} kişi eşleşti ({', '.join(kisiler[:4])}"
                          f"{'…' if len(kisiler) > 4 else ''}) — {len(sirali)} farklı malzeme, "
                          f"{len(bulunan)} hareket")
        else:
            kolonlari_kur(HAREKET_KOLON)
            bulunan.sort(key=lambda r: r[0] or datetime.datetime.min, reverse=True)
            SINIR = 3000
            for i, r in enumerate(bulunan[:SINIR]):
                agac.insert("", "end", tags=("tek",) if i % 2 else (),
                            values=(gt(r[0]), r[1] or "", r[2] or "", r[3] or "",
                                    r[4] if r[4] is not None else "", r[5] or "", r[9]))
            ek = f" (ilk {SINIR} gösteriliyor)" if len(bulunan) > SINIR else ""
            bilgi_var.set(f"{len(kisiler)} kişi eşleşti ({', '.join(kisiler[:4])}"
                          f"{'…' if len(kisiler) > 4 else ''}) — {len(bulunan)} hareket{ek}")

    zamanlayici = [None]

    def geciktir(*_):
        if zamanlayici[0]:
            root.after_cancel(zamanlayici[0])
        zamanlayici[0] = root.after(200, filtrele)

    ad_giris.bind("<KeyRelease>", geciktir)
    malzeme_giris.bind("<KeyRelease>", geciktir)
    islem_cb.bind("<<ComboboxSelected>>", lambda e: filtrele())
    gorunum_cb.bind("<<ComboboxSelected>>", lambda e: filtrele())
    root.bind("<Escape>", lambda e: temizle())

    durum_var = tk.StringVar(value="Veriler yükleniyor…")
    ttk.Label(root, textvariable=durum_var, relief="sunken", anchor="w",
              padding=(10, 4), font=("Segoe UI", 10)).pack(fill="x", side="bottom")

    def islem_listesi_kur():
        turler = sorted({str(r[1]).strip().upper() for r in veri["satirlar"] if r[1]})
        islem_cb["values"] = ["TÜMÜ"] + turler
        if islem_var.get() not in islem_cb["values"]:
            islem_var.set("TÜMÜ")

    def durum_yaz(ek=""):
        satirlar = veri["satirlar"]
        son = max((r[0] for r in satirlar if r[0]), default=None)
        durum_var.set(f"{len(satirlar):,} kayıt".replace(",", ".") +
                      (f" | Son hareket: {son.strftime('%d.%m.%Y')}" if son else "") +
                      (f" | {ek}" if ek else ""))

    def veriyi_al(satirlar, ek):
        veri["satirlar"] = satirlar
        islem_listesi_kur()
        durum_yaz(ek)
        filtrele()

    # önbellekten anında yükle
    onbellek, zaman = cache_oku()
    if onbellek:
        veriyi_al(onbellek, f"önbellek {zaman.strftime('%d.%m %H:%M')} — güncelleniyor…")
    else:
        durum_var.set("İlk kurulum: kaynak dosyalar okunuyor, lütfen bekleyin…")

    def arka_plan():
        try:
            satirlar, hata = tam_guncelle()
            kuyruk.put(("ok", satirlar, hata))
        except Exception as e:
            kuyruk.put(("hata", None, str(e)))

    def kuyruk_kontrol():
        try:
            tip, satirlar, hata = kuyruk.get_nowait()
        except queue.Empty:
            root.after(300, kuyruk_kontrol)
            return
        if tip == "ok":
            veriyi_al(satirlar, "✓ Güncellendi " + datetime.datetime.now().strftime("%H:%M") +
                      (f" — UYARI: {hata}" if hata else ""))
        else:
            if veri["satirlar"]:
                durum_yaz(f"Güncelleme başarısız: {hata} (önbellek gösteriliyor)")
            else:
                messagebox.showerror("Hata", f"Veriler okunamadı:\n{hata}")

    threading.Thread(target=arka_plan, daemon=True).start()
    root.after(300, kuyruk_kontrol)
    root.mainloop()


if __name__ == "__main__":
    if "--rebuild" in sys.argv:
        sys.stdout.reconfigure(encoding="utf-8")
        satirlar, hata = tam_guncelle()
        kaynak = Counter(r[9] for r in satirlar)
        print(f"Toplam {len(satirlar)} kayıt -> {CIKTI_XLSX}")
        for k, v in kaynak.items():
            print(f"  {k}: {v}")
        if hata:
            print("UYARI:", hata)
    else:
        gui()
